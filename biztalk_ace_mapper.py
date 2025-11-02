#!/usr/bin/env python3
"""
BizTalk to ACE Specification-Driven Intelligent Mapper
Focus: Business specification-driven mapping with optimized LLM usage
Author: ACE Migration Expert
"""

import os
import re
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any, Union
import PyPDF2
import groq
from dotenv import load_dotenv
from llm_json_parser import LLMJSONParser
import streamlit as st
import lxml.etree as ET
from vector_knowledge.vector_store import ChromaVectorStore
load_dotenv()

class BizTalkACEMapper:
    """Specification-driven BizTalk to ACE intelligent mapper"""
    
    def __init__(self):
        self.groq_client = None
        self.biztalk_components = []
        self.business_requirements = {}
        self.esql_template = ""
        self.mappings = []
        
        # Initialize LLM client
        self._initialize_llm()
    
    def _initialize_llm(self):
        """Initialize GROQ LLM client"""
        api_key = os.getenv('GROQ_API_KEY')
        if api_key:
            try:
                self.groq_client = groq.Groq(api_key=api_key)
                print("  LLM client initialized")
            except Exception as e:
                print(f"  LLM initialization failed: {e}")
                self.groq_client = None
        else:
            print("  GROQ_API_KEY not found in environment")

    # ========================================
    # RULE-BASED FUNCTIONS (100% Rule-Based)
    # ========================================
    
    def parse_biztalk_components(self, biztalk_files: Any) -> List[Dict]:
        """Parse BizTalk component files from folder scan and extract metadata"""
        components = []
        extensions = {'.btproj', '.odx', '.btm', '.xsl', '.xsd', '.cs'}
        
        # Handle both folder path (string) and file list
        if isinstance(biztalk_files, str):
            # Scan folder recursively (3-4 levels deep)
            folder_path = Path(biztalk_files)
            file_paths = []
            
            for root, dirs, files in os.walk(folder_path):
                # Limit to 4 levels deep
                level = len(Path(root).relative_to(folder_path).parts)
                if level > 4:
                    continue
                
                for file in files:
                    if Path(file).suffix.lower() in extensions:
                        full_path = Path(root) / file
                        file_paths.append(full_path)
            
            print(f"  Scanned folder: Found {len(file_paths)} BizTalk files")
        else:
            # Handle list of file paths
            file_paths = [Path(fp) for fp in biztalk_files]
        
        # Parse each file
        for file_path in file_paths:
            try:
                component = self._parse_single_component(file_path)
                if component:
                    components.append(component)
                    
            except Exception as e:
                print(f"  Failed to parse {file_path.name}: {e}")
                continue
        
        print(f"  Parsed {len(components)} BizTalk components")
        return components



    def detect_flow_pattern(self, vector_db_results: List[Dict]) -> Dict:
        """Detect flow pattern - COMPREHENSIVE DETECTION with Multi-Method Routing"""
        try:
            print("    Analyzing vector DB content...")
            pattern = {
                'input_type': None,
                'has_enrichment': False,
                'has_xsl_transform': False,
                'xsl_files': [],
                'has_soap_request': False,
                'soap_endpoint': None,
                'is_synchronous': False,
                'flow_type': None,
                'methods': [],
                'has_event_nodes': False,
                'node_sequence': [],
                'has_method_routing': False,  # NEW
                'routing_methods': [],         # NEW
                'routing_type': None           # NEW
            }
            
            combined_content = " ".join([r.get('content', '').lower() for r in vector_db_results])
            
            # 1. Input type detection
            if any(kw in combined_content for kw in ['http', 'wcf', 'basichttp', 'web service', 'wsdl', 'soap']):
                pattern['input_type'] = 'HTTP'
                print("      HTTP Input detected")
            elif any(kw in combined_content for kw in ['mq', 'queue', 'mqinput']):
                pattern['input_type'] = 'MQ'
                print("      MQ Input detected")
            else:
                pattern['input_type'] = 'MQ'  # Default
            
            # 2. XSL Transform detection
            import re
            xsl_matches = re.findall(r'([a-zA-Z0-9_\-\.]+\.xsl[t]?)', combined_content, re.IGNORECASE)
            
            if xsl_matches or 'transform' in combined_content or 'xslt' in combined_content or 'mapping' in combined_content:
                pattern['has_xsl_transform'] = True
                pattern['xsl_files'] = list(set(xsl_matches)) if xsl_matches else ['Transform.xsl']
                print(f"      XSL Transform detected: {len(pattern['xsl_files'])} files")
            
            # 3. Enrichment detection
            enrichment_keywords = [
                'enrichment', 'enrich', 'transco', 
                'beforeenrichment', 'afterenrichment',
                'lookup', 'database', 'db lookup',
                'base64', 'encode', 'decode',
                'compression', 'compress'
            ]
            if any(kw in combined_content for kw in enrichment_keywords):
                pattern['has_enrichment'] = True
                print("      Enrichment detected")
            
            # 4. SOAP detection
            if 'soap' in combined_content or 'web service' in combined_content or 'wsdl' in combined_content:
                pattern['has_soap_request'] = True
                
                soap_urls = re.findall(r'https?://[^\s<>"]+', combined_content)
                pattern['soap_endpoint'] = soap_urls[0] if soap_urls else 'http://service.endpoint'
                print(f"      SOAP Request detected: {pattern['soap_endpoint']}")
            
            # 5. Flow type detection
            if any(kw in combined_content for kw in ['rts', 'synchronous', 'request-response', 'two-way', 'reply']):
                pattern['is_synchronous'] = True
                pattern['flow_type'] = 'RTS'
                pattern['has_event_nodes'] = True
                print("      RTS (Synchronous) flow detected")
            elif pattern['input_type'] == 'HTTP':
                pattern['is_synchronous'] = True
                pattern['flow_type'] = 'RTS'
                pattern['has_event_nodes'] = True
                print("      HTTP implies RTS flow")
            
            # 6. Method extraction (NEW)
            method_patterns = [
                r'\b(subscription(?:ws)?)\b',
                r'\b(confirm(?:subscription)?)\b',
                r'\b(cancel(?:subscription|nfs|nfse)?)\b',
                r'\b(submit(?:nfs|nfse)?)\b',
                r'\b([a-z]+(?:ws|service|method))\b'
            ]
            
            found_methods = set()
            for pattern_regex in method_patterns:
                matches = re.findall(pattern_regex, combined_content, re.IGNORECASE)
                found_methods.update([m.lower() for m in matches if len(m) > 2])
            
            pattern['methods'] = list(found_methods)[:10]  # Limit to 10
            
            # 7. Multi-method routing detection (NEW)
            if len(found_methods) >= 3:
                pattern['has_method_routing'] = True
                pattern['routing_methods'] = list(found_methods)
                pattern['routing_type'] = 'route'  # Use Route node pattern
                print(f"      Multi-method routing detected: {len(found_methods)} methods")
                print(f"      Methods: {', '.join(list(found_methods)[:5])}")
            
            # 8. Build node sequence
            pattern['node_sequence'] = self._build_node_sequence(pattern)
            
            print(f"      Pattern Summary:")
            print(f"       Input: {pattern['input_type']}")
            print(f"       Flow Type: {pattern['flow_type']}")
            print(f"       XSL: {pattern['has_xsl_transform']}")
            print(f"       Enrichment: {pattern['has_enrichment']}")
            print(f"       SOAP: {pattern['has_soap_request']}")
            print(f"       Routing: {pattern['has_method_routing']}")
            print(f"       Methods: {len(pattern['methods'])}")
            
            return pattern
            
        except Exception as e:
            print(f"      Pattern detection error: {e}")
            return {
                'input_type': 'MQ',
                'has_enrichment': False,
                'has_xsl_transform': False,
                'xsl_files': [],
                'has_soap_request': False,
                'soap_endpoint': None,
                'is_synchronous': False,
                'flow_type': None,
                'methods': [],
                'has_event_nodes': False,
                'node_sequence': [],
                'has_method_routing': False,
                'routing_methods': [],
                'routing_type': None
            }


    def _build_node_sequence(self, pattern: Dict) -> List[str]:
        """Build complete node sequence for RTS pattern"""
        sequence = []
        
        # Input
        if pattern['input_type'] == 'HTTP':
            sequence.append('HTTPInput')
        else:
            sequence.append('MQInput')
        
        # Event nodes (RTS pattern)
        sequence.append('InputEventMessage')  # ESQL#1
        sequence.append('Compute')             # ESQL#2
        
        # Enrichment
        if pattern['has_enrichment']:
            sequence.append('BeforeEnrichment')
        
        # Transform
        if pattern['has_xsl_transform']:
            sequence.append('XSLTransform')
        
        sequence.append('AfterEnrichment')     # ESQL#3
        
        # SOAP call (if detected)
        if pattern['has_soap_request']:
            sequence.append('SOAPRequest')
        
        sequence.append('OutputEventMessage')  # ESQL#4
        sequence.append('AfterEventMessage')   # ESQL#5
        
        # Output
        if pattern['input_type'] == 'HTTP':
            sequence.append('HTTPReply')
        else:
            sequence.append('MQOutput')
        
        sequence.append('FailureHandler')      # ESQL#6
        
        return sequence
 


    def optimize_msgflow_template(self, vector_db_results: List[Dict], 
                                business_json: Dict = None, 
                                output_path: str = "msgflow_template.xml") -> str:
        """
        Generate optimized MessageFlow template based on business requirements analysis.
        Simplified version with comprehensive logging.
        
        Args:
            vector_db_results: Results from vector DB containing business requirements
            business_json: Business context information (optional)
            output_path: Path to save the generated template
            
        Returns:
            Path to the generated template file
        """
        print("\n========== MSGFLOW TEMPLATE OPTIMIZATION STARTED ==========")
        print(f"Output path: {output_path}")
        
        try:
            # Initialize default business_json if not provided
            if business_json is None:
                business_json = {}
                print("No business_json provided, using empty dict")
            
            # Step 1: Detect the flow pattern from vector DB results
            print("\nSTEP 1: Analyzing business requirements from Vector DB...")
            
            # Simple pattern detection (without helper methods)
            pattern = {
                'input_type': 'MQ',  # Default input type
                'has_enrichment': False,
                'has_xsl_transform': False,
                'has_soap_request': False,
                'is_synchronous': False,
                'flow_type': 'P2P',
                'has_method_routing': False,
                'methods': []
            }
            
            # Combine content from vector_db_results for analysis
            combined_content = ""
            for result in vector_db_results:
                if isinstance(result, dict) and 'content' in result:
                    content = result.get('content', '')
                    if isinstance(content, str):
                        combined_content += content.lower() + " "
            
            if not combined_content:
                print("WARNING: No content found in vector DB results")
            
            # Log the combined content length for debugging
            print(f"Combined content length: {len(combined_content)} characters")
            if len(combined_content) > 100:
                print(f"Content preview: {combined_content[:100]}...")
            
            # Basic pattern detection using keyword search
            print("\nPerforming pattern detection...")
            
            # Input type detection
            if any(kw in combined_content for kw in ['http', 'wcf', 'web service', 'wsdl', 'soap']):
                pattern['input_type'] = 'HTTP'
                print("✓ HTTP Input detected")
            elif 'file input' in combined_content:
                pattern['input_type'] = 'File'
                print("✓ File Input detected")
            else:
                pattern['input_type'] = 'MQ'
                print("✓ MQ Input detected (default)")
            
            # XSL Transform detection
            if any(kw in combined_content for kw in ['transform', 'xslt', 'mapping', '.xsl']):
                pattern['has_xsl_transform'] = True
                print("✓ XSL Transform required")
            else:
                print("× No XSL Transform needed")
            
            # Enrichment detection
            if any(kw in combined_content for kw in ['enrichment', 'enrich', 'lookup', 'database']):
                pattern['has_enrichment'] = True
                print("✓ Enrichment required")
            else:
                print("× No Enrichment needed")
            
            # SOAP Request detection
            if any(kw in combined_content for kw in ['soap', 'web service', 'wsdl']):
                pattern['has_soap_request'] = True
                print("✓ SOAP Request required")
            else:
                print("× No SOAP Request needed")
            
            # Flow type detection
            if pattern['input_type'] == 'HTTP' or 'synchronous' in combined_content:
                pattern['is_synchronous'] = True
                pattern['flow_type'] = 'RTS'
                print("✓ RTS (Synchronous) flow detected")
            else:
                print("✓ P2P (Asynchronous) flow detected")
            
            # Method routing detection
            method_keywords = ['subscription', 'confirm', 'cancel', 'submit', 'method', 'multiple operations']
            if any(kw in combined_content for kw in method_keywords):
                pattern['has_method_routing'] = True
                
                # Extract simple methods
                import re
                methods = []
                method_patterns = [
                    r'\b(subscription(?:ws)?)\b',
                    r'\b(confirm(?:subscription)?)\b',
                    r'\b(cancel(?:subscription|nfs|nfse)?)\b',
                    r'\b(submit(?:nfs|nfse)?)\b'
                ]
                
                for pattern_regex in method_patterns:
                    matches = re.findall(pattern_regex, combined_content)
                    for match in matches:
                        if match and len(match) > 3:  # Filter out very short matches
                            methods.append(match)
                
                if not methods:
                    # Default SAP-Edicom methods
                    methods = ['subscriptionws', 'confirmsubscription', 'cancelnfse', 'submitnfse']
                
                pattern['methods'] = methods
                print(f"✓ Method Routing required with methods: {', '.join(methods[:3])}")
            else:
                print("× No Method Routing needed")
            
            # Log the detected pattern
            print("\nDetected flow pattern:")
            for key, value in pattern.items():
                print(f"  - {key}: {value}")
            
            # Step 2: Load template file
            print("\nSTEP 2: Loading template file...")
            template = None
            template_paths = [
                "templates/messageflow_template_sample.xml",  # Correct path first
                "final_messageflow_template.xml",            # Alternative paths as fallbacks
                "msgflow_template.xml",
                "messageflow_template_sample.xml",
                Path(__file__).parent / "templates" / "messageflow_template_sample.xml"
            ]
            
            template_loaded = False
            for path in template_paths:
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        template = f.read()
                        print(f"✓ Template loaded successfully from: {path}")
                        template_loaded = True
                        break
                except Exception as e:
                    print(f"× Failed to load template from {path}: {e}")
                    continue
            
            if not template_loaded or not template:
                raise ValueError("Failed to load template file from any path")
            
            # Step 3: Process template based on pattern
            print("\nSTEP 3: Processing template based on detected pattern...")
            optimized_template = template
            
            # 3.1. Replace basic placeholders with SELECTIVE {FLOW_NAME} handling
            flow_name = business_json.get('flow_name', 'UnknownFlow')
            app_name = business_json.get('project_name', 'UnknownProject')
            source_queue = business_json.get('source_queue', 'INPUT.QUEUE')
            target_queue = business_json.get('target_queue', 'OUTPUT.QUEUE')
            
            print(f"Replacing basic placeholders with:")
            print(f"  - Flow name: {flow_name}")
            print(f"  - App name: {app_name}")
            print(f"  - Source queue: {source_queue}")
            print(f"  - Target queue: {target_queue}")
            
            # CRITICAL: Protect computeExpression attributes from {FLOW_NAME} replacement
            import re
            
            print("\nProtecting computeExpression attributes from {{FLOW_NAME}} replacement...")
            
            # Extract and store ALL computeExpression values with their indices
            compute_pattern = r'computeExpression="([^"]*)"'
            compute_expressions = []
            
            def extract_compute_expr(match):
                """Extract and store computeExpression values"""
                expr_value = match.group(1)
                if '{FLOW_NAME}' in expr_value:
                    # Store the original expression
                    compute_expressions.append(expr_value)
                    # Replace with a unique placeholder
                    placeholder = f'<<<COMPUTE_EXPR_{len(compute_expressions)-1}>>>'
                    return f'computeExpression="{placeholder}"'
                return match.group(0)  # Return unchanged if no {FLOW_NAME}
            
            # Step 1: Extract all computeExpression attributes containing {FLOW_NAME}
            protected_template = re.sub(compute_pattern, extract_compute_expr, optimized_template)
            
            print(f"  ✓ Protected {len(compute_expressions)} computeExpression attributes")
            if len(compute_expressions) > 0:
                print(f"  ✓ Sample: {compute_expressions[0][:50]}...")
            
            # Step 2: Now safely replace all remaining {FLOW_NAME} placeholders
            protected_template = protected_template.replace('{FLOW_NAME}', flow_name)
            print(f"  ✓ Replaced {{FLOW_NAME}} (except in computeExpression) with: {flow_name}")
            
            # Step 3: Restore the original computeExpression values with preserved {FLOW_NAME}
            def restore_compute_expr(match):
                """Restore original computeExpression values"""
                placeholder = match.group(1)
                if placeholder.startswith('<<<COMPUTE_EXPR_'):
                    # Extract the index
                    idx = int(placeholder.replace('<<<COMPUTE_EXPR_', '').replace('>>>', ''))
                    # Return the original expression
                    return f'computeExpression="{compute_expressions[idx]}"'
                return match.group(0)
            
            restore_pattern = r'computeExpression="([^"]*)"'
            optimized_template = re.sub(restore_pattern, restore_compute_expr, protected_template)
            
            print(f"  ✓ Restored {len(compute_expressions)} computeExpression attributes with {{FLOW_NAME}} preserved")
            
            # Step 4: Replace other placeholders normally
            optimized_template = optimized_template.replace('{APP_NAME}', app_name)
            optimized_template = optimized_template.replace('{INPUT_QUEUE_NAME}', source_queue)
            optimized_template = optimized_template.replace('{OUTPUT_QUEUE_NAME}', target_queue)
            
            # 3.2. Process Input Section
            print("\nProcessing input section...")
            if pattern['input_type'] == 'HTTP':
                input_section = """
        <!-- HTTP Input Node -->
        <nodes xmi:type="ComIbmWSInput.msgnode:FCMComposite_1" 
                xmi:id="FCMComposite_1_7" 
                location="-201,103"
                URLSpecifier="/api/service"
                messageDomainProperty="XMLNSC">
            <translation xmi:type="utility:ConstantString" string="HTTPInput"/>
        </nodes>
                """
                print("Added HTTP Input node")
            elif pattern['input_type'] == 'File':
                input_section = """
        <!-- File Input Node -->
        <nodes xmi:type="ComIbmFileInput.msgnode:FCMComposite_1"
                xmi:id="FCMComposite_1_7"
                location="-201,103"
                inputDirectory="/var/mqsi/input"
                filenamePattern="*.xml"
                messageDomainProperty="XMLNSC">
            <translation xmi:type="utility:ConstantString" string="FileInput"/>
        </nodes>
                """
                print("Added File Input node")
            else:
                input_section = """
        <!-- MQ Input Node -->
        <nodes xmi:type="epis_common_flows_lib_MQInput.subflow:FCMComposite_1" 
                xmi:id="FCMComposite_1_7" 
                location="-201,103">
            <translation xmi:type="utility:ConstantString" string="MQInput"/>
        </nodes>
                """
                print("Added MQ Input node")
                
            optimized_template = optimized_template.replace('{DYNAMIC_INPUT_SECTION}', input_section)
            
            # 3.3. Process Enrichment Section
            print("\nProcessing enrichment section...")
            if pattern['has_enrichment']:
                enrichment_section = """
        <!-- Before Enrichment Node -->
        <nodes xmi:type="epis_enrichment_lib_EPIS_MessageEnrichment.subflow:FCMComposite_1" 
                xmi:id="FCMComposite_1_4" 
                location="30,299" 
                inputDirectory="/var/mqsi/enrichment" 
                filenamePattern="BeforeEnrichmentConf.json">
            <translation xmi:type="utility:ConstantString" string="BeforeEnrichment"/>
        </nodes>
                """
                print("Added Enrichment node")
            else:
                enrichment_section = "<!-- No Enrichment required -->"
                print("Skipped Enrichment node")
                
            optimized_template = optimized_template.replace('{DYNAMIC_ENRICHMENT_SUBFLOW}', enrichment_section)
            
            # 3.4. Process XSL Transform Section
            print("\nProcessing XSL transform section...")
            if pattern['has_xsl_transform']:
                xsl_section = """
        <!-- XSL Transform Node -->
        <nodes xmi:type="ComIbmXslMqsi.msgnode:FCMComposite_1" 
                xmi:id="FCMComposite_1_5" 
                location="596,45" 
                stylesheetName="Transform.xsl" 
                messageDomainProperty="XMLNSC">
            <translation xmi:type="utility:ConstantString" string="XSLTransform"/>
        </nodes>
                """
                print("Added XSL Transform node")
            else:
                xsl_section = "<!-- No XSL Transform required -->"
                print("Skipped XSL Transform node")
                
            optimized_template = optimized_template.replace('{DYNAMIC_XSL_TRANSFORM}', xsl_section)
            
            # 3.5. Process SOAP Section
            print("\nProcessing SOAP section...")
            if pattern['has_soap_request']:
                soap_section = """
        <!-- SOAP Request Node -->
        <nodes xmi:type="ComIbmSOAPRequest.msgnode:FCMComposite_1" 
                xmi:id="FCMComposite_1_10" 
                location="1400,100" 
                wsdlFileName="service.wsdl" 
                selectedPortType="Service" 
                selectedBinding="BasicHttpBinding_Service" 
                selectedOperation="Operation" 
                selectedPort="Port" 
                useHTTPTransport="true" 
                webServiceURL="http://service.endpoint" 
                sslProtocol="TLS">
            <translation xmi:type="utility:ConstantString" string="SOAPRequest"/>
        </nodes>
                """
                print("Added SOAP Request node")
            else:
                soap_section = "<!-- No SOAP Request required -->"
                print("Skipped SOAP Request node")
                
            optimized_template = optimized_template.replace('{DYNAMIC_SOAP_NODE}', soap_section)
            
            # 3.6. Process Method Routing Section
            print("\nProcessing routing section...")
            if pattern['has_method_routing']:
                methods = pattern.get('methods', ['method1', 'method2', 'method3'])
                
                # Create Route node
                route_node = """
        <!-- Route Node -->
        <nodes xmi:type="ComIbmRoute.msgnode:FCMComposite_1" 
                xmi:id="FCMComposite_1_20" 
                location="500,200">
            <translation xmi:type="utility:ConstantString" string="Route"/>
        </nodes>
                """
                
                # Create Label nodes for each method
                label_nodes = ""
                for i, method in enumerate(methods[:3], 1):  # Limit to 3 methods
                    label_nodes += f"""
        <!-- Label Node for {method} -->
        <nodes xmi:type="ComIbmLabel.msgnode:FCMComposite_1" 
                xmi:id="FCMComposite_1_3{i}" 
                location="600,{150 + i*50}" 
                labelName="{method}">
            <translation xmi:type="utility:ConstantString" string="Label_{method}"/>
        </nodes>
                    """
                
                routing_section = route_node + label_nodes
                print(f"Added Routing node with {len(methods[:3])} methods")
            else:
                routing_section = "<!-- No Method Routing required -->"
                print("Skipped Routing node")
                
            optimized_template = optimized_template.replace('{DYNAMIC_ROUTING_SECTION}', routing_section)
            
            # 3.7. Process Output Section
            print("\nProcessing output section...")
            if pattern['is_synchronous'] and pattern['input_type'] == 'HTTP':
                output_section = """
        <!-- HTTP Reply Node -->
        <nodes xmi:type="ComIbmWSReply.msgnode:FCMComposite_1" 
                xmi:id="FCMComposite_1_14" 
                location="1000,100">
            <translation xmi:type="utility:ConstantString" string="HTTPReply"/>
        </nodes>
                """
                print("Added HTTP Reply node")
            else:
                output_section = """
        <!-- MQ Output Node -->
        <nodes xmi:type="epis_common_flows_lib_MQOutput.subflow:FCMComposite_1" 
                xmi:id="FCMComposite_1_14" 
                location="1000,100" 
                NonCDM="_ROUTE_NON_CDM_" 
                STATIC_OUTPUT_QUEUE="{OUTPUT_QUEUE_NAME}" 
                IS_OUTPUT_MQ="false">
            <translation xmi:type="utility:ConstantString" string="MQOutput"/>
        </nodes>
                """
                print("Added MQ Output node")
                
            optimized_template = optimized_template.replace('{DYNAMIC_OUTPUT_SUBFLOW}', output_section)
            
            # 3.8. Generate basic connections
            print("\nGenerating connection patterns...")
            
            connections = """
        <!-- PRIMARY FLOW PATH -->
        
        <!-- Connection 1: MQInput → InputEventMessage -->
        <connections xmi:type="eflow:FCMConnection" xmi:id="FCMConnection_1" 
                    targetNode="FCMComposite_1_2" sourceNode="FCMComposite_1_7" 
                    sourceTerminalName="OutTerminal.Output" targetTerminalName="InTerminal.in"/>

        <!-- Connection 2: InputEventMessage → Compute -->
        <connections xmi:type="eflow:FCMConnection" xmi:id="FCMConnection_2" 
                    targetNode="FCMComposite_1_1" sourceNode="FCMComposite_1_2" 
                    sourceTerminalName="OutTerminal.out" targetTerminalName="InTerminal.in"/>

        <!-- Connection 3: Compute → OutputEventMessage -->
        <connections xmi:type="eflow:FCMConnection" xmi:id="FCMConnection_3" 
                    targetNode="FCMComposite_1_13" sourceNode="FCMComposite_1_1" 
                    sourceTerminalName="OutTerminal.out" targetTerminalName="InTerminal.in"/>
                    
        <!-- Connection 4: OutputEventMessage → MQOutput -->
        <connections xmi:type="eflow:FCMConnection" xmi:id="FCMConnection_4" 
                    targetNode="FCMComposite_1_14" sourceNode="FCMComposite_1_13" 
                    sourceTerminalName="OutTerminal.out" targetTerminalName="InTerminal.Input"/>
                    
        <!-- ERROR HANDLING -->
                    
        <!-- Connection 5: Compute failure → FailureHandler -->
        <connections xmi:type="eflow:FCMConnection" xmi:id="FCMConnection_5" 
                    targetNode="FCMComposite_1_11" sourceNode="FCMComposite_1_1" 
                    sourceTerminalName="OutTerminal.failure" targetTerminalName="InTerminal.in"/>
            """
            
            if pattern['has_soap_request']:
                connections += """
        <!-- SOAP Connections -->
                    
        <!-- Connection 6: OutputEventMessage → SOAPRequest -->
        <connections xmi:type="eflow:FCMConnection" xmi:id="FCMConnection_6" 
                    targetNode="FCMComposite_1_10" sourceNode="FCMComposite_1_13" 
                    sourceTerminalName="OutTerminal.out" targetTerminalName="InTerminal.in"/>
                    
        <!-- Connection 7: SOAPRequest → AfterEventMsg -->
        <connections xmi:type="eflow:FCMConnection" xmi:id="FCMConnection_7" 
                    targetNode="FCMComposite_1_15" sourceNode="FCMComposite_1_10" 
                    sourceTerminalName="OutTerminal.out" targetTerminalName="InTerminal.in"/>
                    
        <!-- Connection 8: SOAPRequest failure → FailureHandler -->
        <connections xmi:type="eflow:FCMConnection" xmi:id="FCMConnection_8" 
                    targetNode="FCMComposite_1_11" sourceNode="FCMComposite_1_10" 
                    sourceTerminalName="OutTerminal.failure" targetTerminalName="InTerminal.in"/>
                """
                print("Added SOAP connections")
                
            optimized_template = optimized_template.replace('{DYNAMIC_CONNECTIONS}', connections)
            
            # 3.9. Generate basic attribute links
            print("\nGenerating attribute links...")
            
            attribute_links = """
        <!-- MQ Input Links -->
        <attributeLinks promotedAttribute="Property.COMPRESS_MSG_SIZE" overriddenNodes="FCMComposite_1_7">
        <overriddenAttribute href="epis/common/flows/lib/MQInput.subflow#Property.COMPRESS_MSG_SIZE"/>
        </attributeLinks>
        
        <attributeLinks promotedAttribute="Property.INPUT_MSG_PARSING" overriddenNodes="FCMComposite_1_7">
        <overriddenAttribute href="epis/common/flows/lib/MQInput.subflow#Property.INPUT_MSG_PARSING"/>
        </attributeLinks>
        
        <attributeLinks promotedAttribute="Property.INPUT_QUEUE" overriddenNodes="FCMComposite_1_7">
        <overriddenAttribute href="epis/common/flows/lib/MQInput.subflow#Property.INPUT_QUEUE"/>
        </attributeLinks>
            """
            
            if pattern['has_soap_request']:
                attribute_links += """
        <!-- SOAP Request Links -->
        <attributeLinks promotedAttribute="Property.SOAP_SERVICE_URL" overriddenNodes="FCMComposite_1_10">
        <overriddenAttribute href="ComIbmSOAPRequest.msgnode#Property.webServiceURL"/>
        </attributeLinks>
        
        <attributeLinks promotedAttribute="Property.WSDL_FILE_NAME" overriddenNodes="FCMComposite_1_10">
        <overriddenAttribute href="ComIbmSOAPRequest.msgnode#Property.wsdlFileName"/>
        </attributeLinks>
                """
                print("Added SOAP attribute links")
                
            optimized_template = optimized_template.replace('{DYNAMIC_ATTRIBUTE_LINKS}', attribute_links)
            
            # 3.10. Process placeholder sections we don't need
            placeholder_sections = [
                '{DYNAMIC_BLOB_STORAGE}',
                '{DYNAMIC_BASE64_ENCODER}',
                '{DYNAMIC_GZIP_SUBFLOW}',
                '{DYNAMIC_HTTP_REQUEST}'
            ]
            
            print("\nReplacing unused placeholder sections...")
            for section in placeholder_sections:
                if section in optimized_template:
                    optimized_template = optimized_template.replace(section, f"<!-- {section[1:-1]} not required -->")
                    print(f"Replaced {section}")
            
            # 3.11. Process conditional properties
            print("\nProcessing conditional properties...")
            import re
            
            # HTTP Input properties
            if pattern['input_type'] == 'HTTP':
                optimized_template = optimized_template.replace('{IF_HTTP_INPUT_START}', '')
                optimized_template = optimized_template.replace('{IF_HTTP_INPUT_END}', '')
                print("Included HTTP Input properties")
            else:
                optimized_template = re.sub(r'{IF_HTTP_INPUT_START}.*?{IF_HTTP_INPUT_END}', '', 
                                        optimized_template, flags=re.DOTALL)
                print("Excluded HTTP Input properties")
            
            # MQ Input properties
            if pattern['input_type'] == 'MQ':
                optimized_template = optimized_template.replace('{IF_MQ_INPUT_START}', '')
                optimized_template = optimized_template.replace('{IF_MQ_INPUT_END}', '')
                print("Included MQ Input properties")
            else:
                optimized_template = re.sub(r'{IF_MQ_INPUT_START}.*?{IF_MQ_INPUT_END}', '', 
                                        optimized_template, flags=re.DOTALL)
                print("Excluded MQ Input properties")
            
            # File Input properties
            if pattern['input_type'] == 'File':
                optimized_template = optimized_template.replace('{IF_FILE_INPUT_START}', '')
                optimized_template = optimized_template.replace('{IF_FILE_INPUT_END}', '')
                print("Included File Input properties")
            else:
                optimized_template = re.sub(r'{IF_FILE_INPUT_START}.*?{IF_FILE_INPUT_END}', '', 
                                        optimized_template, flags=re.DOTALL)
                print("Excluded File Input properties")
            
            # Enrichment properties
            if pattern['has_enrichment']:
                optimized_template = optimized_template.replace('{IF_ENRICHMENT_START}', '')
                optimized_template = optimized_template.replace('{IF_ENRICHMENT_END}', '')
                print("Included Enrichment properties")
            else:
                optimized_template = re.sub(r'{IF_ENRICHMENT_START}.*?{IF_ENRICHMENT_END}', '', 
                                        optimized_template, flags=re.DOTALL)
                print("Excluded Enrichment properties")
            
            # XSL properties
            if pattern['has_xsl_transform']:
                optimized_template = optimized_template.replace('{IF_XSL_START}', '')
                optimized_template = optimized_template.replace('{IF_XSL_END}', '')
                print("Included XSL Transform properties")
            else:
                optimized_template = re.sub(r'{IF_XSL_START}.*?{IF_XSL_END}', '', 
                                        optimized_template, flags=re.DOTALL)
                print("Excluded XSL Transform properties")
            
            # SOAP properties
            if pattern['has_soap_request']:
                optimized_template = optimized_template.replace('{IF_SOAP_START}', '')
                optimized_template = optimized_template.replace('{IF_SOAP_END}', '')
                print("Included SOAP properties")
            else:
                optimized_template = re.sub(r'{IF_SOAP_START}.*?{IF_SOAP_END}', '', 
                                        optimized_template, flags=re.DOTALL)
                print("Excluded SOAP properties")
            
            # Method Routing properties
            if pattern['has_method_routing']:
                optimized_template = optimized_template.replace('{IF_ROUTING_START}', '')
                optimized_template = optimized_template.replace('{IF_ROUTING_END}', '')
                
                # Replace method list with actual methods
                methods = pattern.get('methods', [])
                if methods:
                    method_list = ','.join(methods[:3])  # Limit to 3 methods
                    optimized_template = optimized_template.replace('"method1,method2,method3"', f'"{method_list}"')
                
                print("Included Routing properties")
            else:
                optimized_template = re.sub(r'{IF_ROUTING_START}.*?{IF_ROUTING_END}', '', 
                                        optimized_template, flags=re.DOTALL)
                print("Excluded Routing properties")
            
            # Replace any remaining {NESTED_DESCRIPTORS}
            optimized_template = optimized_template.replace('{NESTED_DESCRIPTORS}', '')
            
            # Step 4: Check for any remaining placeholders
            print("\nSTEP 4: Checking for remaining placeholders...")
            optimized_template = self.clean_conditional_markers(optimized_template)
            import re
            remaining_placeholders = re.findall(r'\{[A-Z_]+\}', optimized_template)
            
            # CRITICAL: Filter out {FLOW_NAME} from remaining placeholders check 
            # as it's intentionally preserved in computeExpression attributes
            remaining_placeholders = [p for p in remaining_placeholders if p != '{FLOW_NAME}']
            
            if remaining_placeholders:
                print(f"WARNING: Found {len(set(remaining_placeholders))} unprocessed placeholders:")
                for placeholder in sorted(set(remaining_placeholders)):
                    print(f"  - {placeholder}")
                
                # Replace remaining placeholders with default values
                print("Replacing remaining placeholders with default values...")
                for placeholder in set(remaining_placeholders):
                    if 'FILE' in placeholder:
                        optimized_template = optimized_template.replace(placeholder, '/var/mqsi/data')
                    elif 'URL' in placeholder or 'SERVICE' in placeholder:
                        optimized_template = optimized_template.replace(placeholder, 'http://service.endpoint')
                    elif 'QUEUE' in placeholder:
                        optimized_template = optimized_template.replace(placeholder, 'QUEUE.NAME')
                    else:
                        optimized_template = optimized_template.replace(placeholder, placeholder[1:-1])
            else:
                print("✓ No remaining placeholders found (except {FLOW_NAME} in computeExpression)")
            
            # Step 5: Validate and save template
            print("\nSTEP 5: Validating and saving template...")
            
            # Basic validation - check for required sections
            required_elements = [
                '<ecore:EPackage',
                '<eClassifiers',
                '<composition>',
                '</composition>',
                '<propertyOrganizer>',
                '</propertyOrganizer>',
                '<stickyBoard/>',
                '</ecore:EPackage>'
            ]
            
            missing_elements = []
            for element in required_elements:
                if element not in optimized_template:
                    missing_elements.append(element)
            
            if missing_elements:
                print(f"WARNING: Missing required elements: {missing_elements}")
            else:
                print("✓ All required elements present")
            
            # Basic XML validation
            try:
                import xml.dom.minidom as minidom
                minidom.parseString(optimized_template)
                print("✓ XML validation successful")
            except Exception as e:
                print(f"WARNING: XML validation failed: {e}")
                print("Continuing despite XML validation failure")
            
            # Create parent directory if needed
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            
            # Write optimized template to file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(optimized_template)
                
            print(f"\n✓ Successfully saved optimized template to: {output_path}")
            print(f"Template size: {len(optimized_template)} characters")
            print("\n========== MSGFLOW TEMPLATE OPTIMIZATION COMPLETED ==========")
            
            return output_path
            
        except Exception as e:
            # Detailed error reporting with traceback
            import traceback
            error_details = traceback.format_exc()
            
            print(f"\n❌ ERROR: Template optimization failed: {str(e)}")
            print("\nDetailed error traceback:")
            print(error_details)
            
            # Create parent directory if needed
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            
            # Write error information to a log file
            log_path = output_path + ".error.log"
            with open(log_path, 'w', encoding='utf-8') as f:
                f.write(f"Error: {str(e)}\n\n")
                f.write("Traceback:\n")
                f.write(error_details)
                
            print(f"\nError details written to: {log_path}")
            print("\n========== MSGFLOW TEMPLATE OPTIMIZATION FAILED ==========")
            
            # Return the error log path
            return log_path


    def clean_conditional_markers(self, template: str) -> str:
        """
        Clean up all conditional markers and stray comments from the template
        
        Args:
            template: The template string to clean
            
        Returns:
            Cleaned template string
        """
        import re
        
        print("\nCleaning up conditional markers and stray comments...")
        
        # 1. Remove any remaining conditional markers
        conditional_markers = [
            'IF_HTTP_INPUT_START', 'IF_HTTP_INPUT_END',
            'IF_MQ_INPUT_START', 'IF_MQ_INPUT_END',
            'IF_FILE_INPUT_START', 'IF_FILE_INPUT_END',
            'IF_BLOB_STORAGE_START', 'IF_BLOB_STORAGE_END',
            'IF_ENRICHMENT_START', 'IF_ENRICHMENT_END',
            'IF_XSL_START', 'IF_XSL_END',
            'IF_SOAP_START', 'IF_SOAP_END',
            'IF_HTTP_REQUEST_START', 'IF_HTTP_REQUEST_END',
            'IF_MQ_OUTPUT_START', 'IF_MQ_OUTPUT_END',
            'IF_FILE_OUTPUT_START', 'IF_FILE_OUTPUT_END',
            'IF_ROUTING_START', 'IF_ROUTING_END'
        ]
        
        cleaned_template = template
        
        # Count remaining markers for logging
        total_markers_removed = 0
        
        # First, check how many markers exist in the template
        for marker in conditional_markers:
            marker_count = cleaned_template.count(marker)
            if marker_count > 0:
                total_markers_removed += marker_count
        
        if total_markers_removed > 0:
            print(f"  - Found {total_markers_removed} conditional markers to remove")
        
        # Remove markers one by one
        for marker in conditional_markers:
            original_length = len(cleaned_template)
            cleaned_template = cleaned_template.replace('{' + marker + '}', '')
            new_length = len(cleaned_template)
            if original_length != new_length:
                print(f"  - Removed marker: {marker}")
        
        # 2. Remove stray file path comments (like /var/mqsi/data)
        # Use regex to find isolated filepath lines that aren't part of XML tags
        filepath_pattern = r'\n\s*/var/mqsi/\w+\s*\n'
        filepath_matches = re.findall(filepath_pattern, cleaned_template)
        
        if filepath_matches:
            print(f"  - Found {len(filepath_matches)} stray filepath comments")
            cleaned_template = re.sub(filepath_pattern, '\n', cleaned_template)
        
        # 3. Remove any XML comments that contain placeholder documentation
        comment_patterns = [
            r'<!--\s*Possible inputs.*?-->',
            r'<!--\s*[A-Za-z]+ node.*?-->',
            r'<!--\s*No\s+\w+\s+required\s*-->',
            r'<!--\s*DYNAMIC_[A-Z_]+\s+not\s+required\s*-->'
        ]
        
        for pattern in comment_patterns:
            matches = re.findall(pattern, cleaned_template, re.DOTALL)
            if matches:
                print(f"  - Found {len(matches)} documentation comments")
                cleaned_template = re.sub(pattern, '', cleaned_template, flags=re.DOTALL)
        
        # 4. Remove any NESTED_DESCRIPTORS placeholders
        nested_descriptor_count = cleaned_template.count('{NESTED_DESCRIPTORS}')
        if nested_descriptor_count > 0:
            cleaned_template = cleaned_template.replace('{NESTED_DESCRIPTORS}', '')
            print(f"  - Removed {nested_descriptor_count} NESTED_DESCRIPTORS placeholders")
        
        print("✅ Template cleanup complete")
        return cleaned_template


    


    def _parse_single_component(self, file_path: Path) -> Optional[Dict]:
        """Parse individual BizTalk component file"""
        extension = file_path.suffix.lower()
        
        component = {
            'name': file_path.stem,
            'file_name': file_path.name,
            'file_type': extension,
            'path': str(file_path),
            'size': file_path.stat().st_size if file_path.exists() else 0
        }
        
        # Rule-based component type classification
        if extension == '.btproj':
            component['type'] = 'BizTalk Project'
            component['complexity'] = 'LOW'
        elif extension == '.odx':
            component['type'] = 'Orchestration'
            component['complexity'] = 'HIGH'
        elif extension == '.btm':
            component['type'] = 'BizTalk Map'
            component['complexity'] = 'MEDIUM'
        elif extension == '.xsl':
            component['type'] = 'XSLT Transform'
            component['complexity'] = 'MEDIUM'
        elif extension == '.xsd':
            component['type'] = 'Schema Definition'
            component['complexity'] = 'LOW'
        elif extension == '.cs':
            component['type'] = 'C# Code'
            component['complexity'] = 'HIGH'
        else:
            component['type'] = 'Other File'
            component['complexity'] = 'UNKNOWN'
        
        return component
    
    def extract_pdf_text(self, pdf_file: str) -> str:
        """Extract text content from PDF file"""
        try:
            with open(pdf_file, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text_content = ""
                
                for page in pdf_reader.pages:
                    text_content += page.extract_text() + "\n"
                
                print(f"  Extracted {len(text_content)} characters from PDF")
                return text_content.strip()
                
        except Exception as e:
            print(f"  PDF extraction failed: {e}")
            return ""
    
    def load_standard_esql_template(self) -> str:
        """Load the standard ESQL template"""
        template = """
CREATE COMPUTE MODULE _SYSTEM___MSG_TYPE___FLOW_PROCESS___SYSTEM2___FLOW_TYPE__InputEventMessage
	CREATE FUNCTION Main() RETURNS BOOLEAN
	BEGIN
		
		DECLARE episInfo 		REFERENCE TO 	Environment.variables.EventData.episInfo;
		DECLARE sourceInfo 		REFERENCE TO 	Environment.variables.EventData.sourceInfo;
		DECLARE targetInfo 		REFERENCE TO 	Environment.variables.EventData.targetInfo;
		DECLARE integrationInfo REFERENCE TO 	Environment.variables.EventData.integrationInfo;
		DECLARE dataInfo 		REFERENCE TO 	Environment.variables.EventData.dataInfo;
		
		SET sourceInfo.srcAppIdentifier 		= InputRoot.XMLNSC.[<].*:Header.*:Source.*:Identifier; 
		SET sourceInfo.srcEnterpriseCode	 	= InputRoot.XMLNSC.[<].*:Header.*:Source.*:EnterpriseCode;
		SET sourceInfo.srcDivision		 		= InputRoot.XMLNSC.[<].*:Header.*:Source.*:Division;
		SET sourceInfo.srcDepartmentCode 		= InputRoot.XMLNSC.[<].*:Header.*:Source.*:DepartmentCode;
		SET sourceInfo.srcBranchCode 			= InputRoot.XMLNSC.[<].*:Header.*:Source.*:BranchCode;
		SET sourceInfo.srcCountryCode 			= InputRoot.XMLNSC.[<].*:Header.*:Source.*:CountryCode;	
		SET sourceInfo.srcCompanyCode 			= InputRoot.XMLNSC.[<].*:Header.*:Source.*:CompanyCode;
		SET sourceInfo.srcApplicationCode 		= InputRoot.XMLNSC.[<].*:Header.*:Source.*:ApplicationCode;
		
		SET targetInfo.tgtAppIdentifier 		= InputRoot.XMLNSC.[<].*:Header.*:Target.*:Identifier; 	
		SET targetInfo.tgtEnterpriseCode 		= InputRoot.XMLNSC.[<].*:Header.*:Target.*:EnterpriseCode; 
		SET targetInfo.tgtDivision 				= InputRoot.XMLNSC.[<].*:Header.*:Target.*:Division; 
		SET targetInfo.tgtDepartmentCode 		= InputRoot.XMLNSC.[<].*:Header.*:Target.*:DepartmentCode; 
		SET targetInfo.tgtBranchCode 			= InputRoot.XMLNSC.[<].*:Header.*:Target.*:branchCode;
		SET targetInfo.tgtCountryCode 			= InputRoot.XMLNSC.[<].*:Header.*:Target.*:CountryCode;  
		SET targetInfo.tgtCompanyCode 			= InputRoot.XMLNSC.[<].*:Header.*:Target.*:CompanyCode; 
		SET targetInfo.tgtApplicationCode 		= InputRoot.XMLNSC.[<].*:Header.*:Target.*:ApplicationCode; 
	
		SET dataInfo.messageType = InputRoot.XMLNSC.[<].*:Header.*:MessageType;		
		SET dataInfo.dataFormat = 'XML';
		SET dataInfo.mainIdentifier = InputRoot.XMLNSC.[<].*:ShipmentInstruction.*:ShipmentDetails.*:ShipmentId;
		SET dataInfo.customReference1		= ''; 						
		SET dataInfo.customReference1Type	= ''; 	
		SET dataInfo.customReference2		= ''; 	
		SET dataInfo.customReference2Type	= ''; 	
		SET dataInfo.customReference3		= ''; 	
		SET dataInfo.customReference3Type	= ''; 
		SET dataInfo.customReference4		= ''; 
		SET dataInfo.customReference4Type	= '';
		SET dataInfo.customProperty1		= '';
		SET dataInfo.customProperty1Type	= '';
		SET dataInfo.customProperty2		= '';
		SET dataInfo.customProperty2Type	= '';
		SET dataInfo.customProperty3		= '';
		SET dataInfo.customProperty3Type	= '';
		SET dataInfo.customProperty4		= '';
		SET dataInfo.customProperty4Type	= '';
		SET dataInfo.batch = false;
		SET OutputRoot=NULL;

	RETURN TRUE;
	END;
END MODULE;
"""
        self.esql_template = template.strip()
        print("Standard ESQL template loaded")
        return self.esql_template
    
    def validate_esql_syntax(self, esql_content: str) -> Dict:
        """Validate ESQL template syntax"""
        validation = {
            'valid': True,
            'errors': [],
            'warnings': []
        }
        
        # Basic syntax validation
        required_elements = [
            'CREATE COMPUTE MODULE',
            'CREATE FUNCTION Main()',
            'RETURNS BOOLEAN',
            'BEGIN',
            'END;',
            'END MODULE;'
        ]
        
        for element in required_elements:
            if element not in esql_content:
                validation['errors'].append(f"Missing required element: {element}")
                validation['valid'] = False
        
        # Check for forbidden patterns
        forbidden_patterns = ['@', '^esql', '```']
        for pattern in forbidden_patterns:
            if pattern in esql_content:
                validation['errors'].append(f"Forbidden pattern found: {pattern}")
                validation['valid'] = False
        
        print(f"ESQL validation: {'Valid' if validation['valid'] else 'Invalid'}")
        return validation
    
    # ========================================
    # LLM-BASED FUNCTIONS (100% LLM-Based)
    # ========================================
    


    def generate_excel_output(self, mappings: List[Dict], output_path: str) -> str:
        """Generate comprehensive Excel file with multiple sheets for detailed component mapping"""
        try:
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # SHEET 1: Component Overview (Main Summary)
                overview_data = []
                for mapping in mappings:
                    overview_data.append({
                        'BizTalk Component': mapping.get('biztalk_component', ''),
                        'Component Type': mapping.get('component_type', ''),
                        'Business Purpose': mapping.get('business_purpose', ''),
                        'Primary ACE Artifact': mapping.get('ace_components', {}).get('primary_artifact', {}).get('name', ''),
                        'Primary Artifact Type': mapping.get('ace_components', {}).get('primary_artifact', {}).get('type', ''),
                        'Implementation Priority': mapping.get('implementation_priority', ''),
                        'Confidence': mapping.get('confidence', 0),
                        'Input Queue': mapping.get('ace_components', {}).get('integration_details', {}).get('input_queue', ''),
                        'Output Endpoint': mapping.get('ace_components', {}).get('integration_details', {}).get('output_endpoint', ''),
                        'Reasoning': mapping.get('reasoning', '')
                    })
                
                overview_df = pd.DataFrame(overview_data)
                overview_df.to_excel(writer, sheet_name='Component Overview', index=False)
                
                # SHEET 2: ACE Artifacts Detail
                artifacts_data = []
                for mapping in mappings:
                    biztalk_comp = mapping.get('biztalk_component', '')
                    
                    # Primary artifact
                    primary = mapping.get('ace_components', {}).get('primary_artifact', {})
                    if primary:
                        artifacts_data.append({
                            'BizTalk Component': biztalk_comp,
                            'Artifact Category': 'Primary',
                            'Artifact Type': primary.get('type', ''),
                            'Artifact Name': primary.get('name', ''),
                            'Description/Purpose': primary.get('description', ''),
                            'Implementation Notes': f"Main {primary.get('type', '')} for {biztalk_comp}"
                        })
                    
                    # Supporting artifacts
                    supporting = mapping.get('ace_components', {}).get('supporting_artifacts', [])
                    for artifact in supporting:
                        artifacts_data.append({
                            'BizTalk Component': biztalk_comp,
                            'Artifact Category': 'Supporting',
                            'Artifact Type': artifact.get('type', ''),
                            'Artifact Name': artifact.get('name', ''),
                            'Description/Purpose': artifact.get('purpose', ''),
                            'Implementation Notes': f"Supporting {artifact.get('type', '')} for business logic"
                        })
                
                artifacts_df = pd.DataFrame(artifacts_data)
                artifacts_df.to_excel(writer, sheet_name='ACE Artifacts', index=False)
                
                # SHEET 3: Database Operations
                database_data = []
                for mapping in mappings:
                    biztalk_comp = mapping.get('biztalk_component', '')
                    db_operations = mapping.get('ace_components', {}).get('integration_details', {}).get('database_operations', [])
                    
                    for db_op in db_operations:
                        database_data.append({
                            'BizTalk Component': biztalk_comp,
                            'Database Operation': db_op,
                            'Operation Type': 'Stored Procedure' if db_op.startswith('sp_') or db_op.startswith('proc_') else 'Database Call',
                            'Purpose': f"Data lookup/enrichment for {biztalk_comp}",
                            'Implementation Priority': mapping.get('implementation_priority', 'medium')
                        })
                
                database_df = pd.DataFrame(database_data)
                if not database_df.empty:
                    database_df.to_excel(writer, sheet_name='Database Operations', index=False)
                
                # SHEET 4: Integration Endpoints
                integration_data = []
                for mapping in mappings:
                    biztalk_comp = mapping.get('biztalk_component', '')
                    integration_details = mapping.get('ace_components', {}).get('integration_details', {})
                    
                    input_queue = integration_details.get('input_queue')
                    output_endpoint = integration_details.get('output_endpoint')
                    transformation_logic = integration_details.get('transformation_logic', '')
                    
                    if input_queue or output_endpoint:
                        integration_data.append({
                            'BizTalk Component': biztalk_comp,
                            'Input Queue': input_queue or 'N/A',
                            'Output Endpoint': output_endpoint or 'N/A',
                            'Transformation Logic': transformation_logic,
                            'Integration Pattern': 'Queue-to-Service' if input_queue and output_endpoint else 'Simple Transform',
                            'Configuration Notes': f"Configure MQ settings for {input_queue}" if input_queue else "Internal transformation only"
                        })
                
                integration_df = pd.DataFrame(integration_data)
                if not integration_df.empty:
                    integration_df.to_excel(writer, sheet_name='Integration Points', index=False)
                
                # SHEET 5: Implementation Roadmap
                roadmap_data = []
                priority_order = {'high': 1, 'medium': 2, 'low': 3}
                
                for mapping in mappings:
                    biztalk_comp = mapping.get('biztalk_component', '')
                    priority = mapping.get('implementation_priority', 'medium')
                    primary_artifact = mapping.get('ace_components', {}).get('primary_artifact', {})
                    supporting_artifacts = mapping.get('ace_components', {}).get('supporting_artifacts', [])
                    
                    total_artifacts = 1 + len(supporting_artifacts)  # Primary + supporting
                    estimated_effort = 'High' if total_artifacts > 3 else 'Medium' if total_artifacts > 1 else 'Low'
                    
                    roadmap_data.append({
                        'Priority Rank': priority_order.get(priority, 2),
                        'BizTalk Component': biztalk_comp,
                        'Implementation Priority': priority.title(),
                        'Primary Deliverable': primary_artifact.get('name', ''),
                        'Total Artifacts': total_artifacts,
                        'Estimated Effort': estimated_effort,
                        'Dependencies': ', '.join([art.get('name', '') for art in supporting_artifacts[:3]]),  # First 3 dependencies
                        'Business Impact': mapping.get('business_purpose', '')[:100] + '...' if len(mapping.get('business_purpose', '')) > 100 else mapping.get('business_purpose', '')
                    })
                
                # Sort by priority
                roadmap_df = pd.DataFrame(roadmap_data)
                roadmap_df = roadmap_df.sort_values('Priority Rank').drop('Priority Rank', axis=1)
                roadmap_df.to_excel(writer, sheet_name='Implementation Roadmap', index=False)
                
                # SHEET 6: Legacy Format (for backward compatibility)
                legacy_data = []
                for mapping in mappings:
                    legacy_data.append({
                        'biztalk_component': mapping.get('biztalk_component', ''),
                        'required_ace_library': f"Component-Level: {mapping.get('ace_components', {}).get('primary_artifact', {}).get('type', 'unknown')}"
                    })
                
                legacy_df = pd.DataFrame(legacy_data)
                legacy_df.to_excel(writer, sheet_name='Component Mapping', index=False)
            
            # Format Excel file for better readability
            self._format_excel_file(output_path)
            
            print(f"Enhanced Excel file generated: {output_path}")
            print(f"Component mappings: {len(mappings)}")
            print(f"Excel sheets: 6 (Overview, Artifacts, Database, Integration, Roadmap, Legacy)")
            
            return output_path
            
        except Exception as e:
            print(f"Excel generation failed: {e}")
            raise

    def _format_excel_file(self, file_path: str):
        """Apply formatting to Excel file for better readability"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            workbook = load_workbook(file_path)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Format headers (first row)
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(file_path)
            print(f"Excel formatting applied")
            
        except Exception as e:
            print(f"Excel formatting failed (but file still usable): {e}")

    def generate_intelligent_mappings(self, biztalk_components: List[Dict], business_requirements: Dict) -> List[Dict]:
        """Generate intelligent BizTalk to ACE component mappings based on business specifications"""
        
        # Validate prerequisites
        if not self.groq_client:
            raise Exception("LLM client required for component-level mapping. Please configure GROQ_API_KEY.")
        
        if not business_requirements:
            raise Exception("Business requirements required for intelligent mapping. Please provide PDF documentation.")
        
        if not biztalk_components:
            raise Exception("No BizTalk components found for mapping.")
        
        try:
            # Prepare component summary
            component_summary = []
            for comp in biztalk_components[:10]:  # Limit to first 10 components
                component_summary.append({
                    'name': comp.get('name', ''),
                    'type': comp.get('type', ''),
                    'complexity': comp.get('complexity', '')
                })
            
            # Create LLM prompt for component-level mapping
            prompt = f"""As an IBM ACE expert, create detailed component mappings for BizTalk to ACE migration.

    BIZTALK COMPONENTS TO MAP:
    {json.dumps(component_summary, indent=2)}

    BUSINESS REQUIREMENTS CONTEXT:
    {json.dumps(business_requirements, indent=2)[:2000]}

    For each BizTalk component, generate specific ACE artifacts with this structure:

    [
        {{
            "biztalk_component": "exact component name",
            "component_type": "BizTalk component type",
            "business_purpose": "what this component does in business terms",
            "ace_components": {{
                "primary_artifact": {{
                    "type": "message_flow|compute_node|xsl_transform|esql_module",
                    "name": "specific_file_name.msgflow|.esql|.xsl",
                    "description": "purpose of this primary artifact"
                }},
                "supporting_artifacts": [
                    {{
                        "type": "esql_module|xsl_transform|subflow|database_lookup",
                        "name": "specific_file_name.ext",
                        "purpose": "what this supporting artifact does"
                    }}
                ],
                "integration_details": {{
                    "input_queue": "specific_queue_name_if_applicable",
                    "output_endpoint": "service_or_queue_name_if_applicable",
                    "database_operations": ["stored_procedure_names"],
                    "transformation_logic": "brief_description_of_business_rules"
                }}
            }},
            "implementation_priority": "high|medium|low",
            "confidence": 0.85,
            "reasoning": "why this component mapping makes business sense"
        }}
    ]

    REQUIREMENTS:
    1. Generate specific, actionable ACE component names
    2. Include database operations from business requirements
    3. Map integration endpoints (queues, services)
    4. Provide transformation logic based on business context
    5. Return valid JSON structure"""

            # Call LLM
            response = self.groq_client.chat.completions.create(
                model=os.getenv('GROQ_MODEL', 'llama-3.1-8b-instant'),
                messages=[
                    {
                        "role": "system", 
                        "content": "You are an expert IBM ACE architect specializing in detailed component-level BizTalk to ACE migrations. Generate precise, actionable component mappings."
                    },
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,  # Low temperature for consistent output
                max_tokens=4000   # Higher limit for detailed component mappings
            )
            if 'token_tracker' in st.session_state and hasattr(response, 'usage') and response.usage:
                st.session_state.token_tracker.manual_track(
                    agent="biztalk_mapper",
                    operation="component_mapping",
                    model=os.getenv('GROQ_MODEL', 'llama-3.1-8b-instant'),
                    input_tokens=response.usage.prompt_tokens,
                    output_tokens=response.usage.completion_tokens,
                    flow_name=getattr(self, 'current_flow_name', 'component_mapping')
                )
            
            # Parse LLM response
            mappings = self.json_parser.parse_component_mappings(response.choices[0].message.content)
            
            # Validate output quality
            if not mappings:
                raise Exception("LLM generated no valid mappings. Check prompt or parsing logic.")
            
            # Quality check - ensure primary artifacts are specified
            valid_mappings = []
            for mapping in mappings:
                if (mapping.get('ace_components') and 
                    mapping.get('ace_components').get('primary_artifact') and
                    mapping.get('ace_components').get('primary_artifact').get('name')):
                    valid_mappings.append(mapping)
                else:
                    print(f"Skipping incomplete mapping for: {mapping.get('biztalk_component', 'Unknown')}")
            
            if not valid_mappings:
                raise Exception("No valid component mappings generated. All mappings failed quality validation.")
            
            print(f"Generated {len(valid_mappings)} intelligent component mappings")
            return valid_mappings
            
        except Exception as e:
            raise Exception(f"Component mapping generation failed: {str(e)}")

    def save_mapping_outputs(self, mappings: List[Dict], output_dir: str) -> Dict[str, str]:
        """Save both JSON and Excel outputs for component mappings"""
        
        try:
            os.makedirs(output_dir, exist_ok=True)
            
            # OUTPUT 1: JSON file for successive programs
            json_file_path = os.path.join(output_dir, "biztalk_ace_component_mapping.json")
            
            # Add metadata to JSON output   
            json_output = {
                "metadata": {
                    "generated_at": datetime.now().isoformat(),
                    "mapping_type": "component_level",
                    "total_components": len(mappings),
                    "generator": "BizTalkACEMapper",
                    "version": "2.0"
                },
                "component_mappings": mappings
            }
            
            with open(json_file_path, 'w', encoding='utf-8') as f:
                json.dump(json_output, f, indent=2, ensure_ascii=False)
            
            print(f"JSON file saved: {json_file_path}")
            print(f"   Input for Programs 3, 4, 5...")
            
            # OUTPUT 2: Excel file for user documentation  
            excel_file_path = os.path.join(output_dir, "biztalk_ace_component_mapping.xlsx")
            excel_path = self.generate_excel_output(mappings, excel_file_path)
            
            print(f"Excel file saved: {excel_path}")
            print(f"   Documentation and user review")
            
            return {
                "json_file": json_file_path,
                "excel_file": excel_path
            }
            
        except Exception as e:
            raise Exception(f"Failed to save mapping outputs: {str(e)}")



    def process_mapping(self, biztalk_files: Union[str, List[str], None], pdf_file: str, output_dir: str) -> Dict:
        """Main processing function with Vector DB integration - Supports PDF-only mode"""
        
        try:
            print("Starting Specification-Driven Component Mapping")
            print("=" * 60)
            
            # Phase 1: BizTalk Component Analysis (Rule-Based) - OPTIONAL
            print("Phase 1: Analyzing BizTalk components...")
            
            if biztalk_files:
                self.biztalk_components = self.parse_biztalk_components(biztalk_files)
                
                if not self.biztalk_components:
                    print("No BizTalk components found in provided path")
                    self.biztalk_components = []
                else:
                    print(f"Found {len(self.biztalk_components)} BizTalk components")
            else:
                print("No BizTalk path provided - PDF-only mode enabled")
                self.biztalk_components = []
            
            # Phase 2: Business Requirements Extraction - VECTOR DB ONLY
            print("Phase 2: Processing Vector DB focused content...")
            
            if not pdf_file:
                raise Exception("Vector DB Error: No focused content received from Vector search")
            
            # Use Vector DB content directly
            print(f"Vector content received: {len(pdf_file)} characters")
            self.business_requirements = self.extract_business_requirements(pdf_file)
            
            if not self.business_requirements:
                raise Exception("Vector DB Error: Failed to extract business requirements from focused content")
            
            print(f"Vector business requirements processed successfully")
            
            # NEW PHASE: Generate MessageFlow Template based on business requirements
            print("Phase 2.5: Generating MessageFlow template...")
            
            # Prepare vector_db_results from pdf_file content
            vector_db_results = [
                {'content': pdf_file}
            ]
            
            # Prepare business_json from business requirements
            business_json = {
                'flow_name': self.business_requirements.get('message_flows', ['UnknownFlow'])[0] if self.business_requirements.get('message_flows') else 'UnknownFlow',
                'project_name': self.business_requirements.get('ace_library_indicators', ['UnknownProject'])[0] if self.business_requirements.get('ace_library_indicators') else 'UnknownProject',
                'source_queue': self.business_requirements.get('integration_endpoints', ['INPUT.QUEUE'])[0] if self.business_requirements.get('integration_endpoints') else 'INPUT.QUEUE',
                'target_queue': 'OUTPUT.QUEUE',
                'properties': [],
                'structural_features': []
            }
            
            print(f"   Flow: {business_json['flow_name']}")
            print(f"   Project: {business_json['project_name']}")
            
            # Generate MessageFlow template
            msgflow_path = None
            try:
                msgflow_path = self.optimize_msgflow_template(
                    vector_db_results,
                    business_json,
                    output_path="msgflow_template.xml"
                )
                print(f"MessageFlow template generated: {msgflow_path}")
            except Exception as e:
                print(f"MessageFlow template generation failed: {str(e)}")
                import traceback
                print(f"   Traceback:\n{traceback.format_exc()}")
            
            # Phase 3: Component-Level Mapping (LLM) - CONDITIONAL
            print("Phase 3: Generating component mappings...")
            
            #   FIX: Only generate mappings if BizTalk components exist
            if self.biztalk_components:
                self.mappings = self.generate_intelligent_mappings(
                    self.biztalk_components,
                    self.business_requirements  
                )
                
                if not self.mappings:
                    print("No valid component mappings generated")
                    self.mappings = []
                else:
                    print(f"Generated {len(self.mappings)} component mappings")
            else:
                print("Skipping component mappings (No BizTalk components to map)")
                self.mappings = []
            
            # Phase 4: ESQL Template Customization (Optional Enhancement)
            print("Phase 4: Customizing ESQL template...")
            
            customized_path = None
            #   FIX: Check if template_path exists and is valid
            template_path = "templates/messageflow_template_sample_v2.xml"
            
            if template_path and os.path.exists(template_path):
                try:
                    with open(template_path, 'r', encoding='utf-8') as f:
                        self.esql_template = f.read()
                    
                    # Customize template using mappings (works even without BizTalk components)
                    customized_template = self.customize_esql_template(
                        self.esql_template, 
                        self.biztalk_components, 
                        self.business_requirements
                    )
                    
                    # Save customized template
                    customized_path = os.path.join(output_dir, "ESQL_Template_Updated.esql")
                    os.makedirs(output_dir, exist_ok=True)
                    with open(customized_path, 'w', encoding='utf-8') as f:
                        f.write(customized_template)
                    print(f"Customized ESQL template saved: {customized_path}")
                except Exception as e:
                    print(f"ESQL template customization failed: {e}")
                    customized_path = None
            else:
                print("Base ESQL template not found, skipping customization")
            
            # Phase 5: Generate Output Files - CONDITIONAL
            print("Phase 5: Generating output files...")
            
            # Initialize output_files dictionary
            output_files = {"json_file": None, "excel_file": None}
            
            # Only generate mapping outputs if we have mappings
            if self.mappings:
                output_files = self.save_mapping_outputs(self.mappings, output_dir)
                print(f"Mapping outputs saved")
            else:
                print("No component mappings to save (PDF-only mode)")
                # Still save business requirements for downstream use
                requirements_path = os.path.join(output_dir, "business_requirements.json")
                os.makedirs(output_dir, exist_ok=True)
                with open(requirements_path, 'w', encoding='utf-8') as f:
                    json.dump(self.business_requirements, f, indent=2, ensure_ascii=False)
                print(f"Business requirements saved: {requirements_path}")
                output_files["json_file"] = requirements_path
            
            # Return comprehensive results
            result = {
                "components_processed": len(self.biztalk_components),
                "mappings_generated": len(self.mappings),
                "business_requirements_found": True,
                "json_file": output_files.get("json_file"),
                "excel_file": output_files.get("excel_file"),
                "template_file": customized_path,
                "vector_processing": True,
                "vector_content_length": len(pdf_file),
                "mode": "pdf_only" if not self.biztalk_components else "full_mapping"
            }
            
            # Add MessageFlow template path to results if it was created
            if msgflow_path:
                result["msgflow_template"] = msgflow_path
            
            print("  Specification-driven mapping completed successfully!")
            if not self.biztalk_components:
                print("  PDF-only mode: Business requirements extracted, ready for Agent 2")
            else:
                print(f"  Full mapping mode: {len(self.mappings)} components mapped")
            
            return result
            
        except Exception as e:
            # Let Vector DB errors propagate to UI
            error_message = str(e)
            if "Vector DB Error" in error_message:
                print(f"  {error_message}")
                raise Exception(f"Vector DB Processing Failed: {error_message}")
            else:
                print(f"  Error in process_mapping: {error_message}")
                raise Exception(f"Component Mapping Failed: {error_message}")
        


    def extract_business_requirements(self, pdf_content: str) -> Dict:
        """
        Extract comprehensive business requirements for all messageflows
        using the template-driven approach with enhanced error handling
        
        Args:
            pdf_content: PDF content text
            
        Returns:
            Dict with extracted business requirements according to template
        """
        try:
            print("🔍 Extracting comprehensive business requirements...")
            
            # CRITICAL FIX: Create a safer wrapper to access vector DB
            def safe_vector_query(query_text):
                try:
                    return self._query_vector_db(query_text)
                except Exception as e:
                    print(f"  ⚠️ Vector query failed: {str(e)}")
                    return []
            
            # Step 1: Discover applications via naming conventions
            naming_files = self._discover_naming_conventions()
            print(f"📋 Found {len(naming_files)} naming convention files")
            
            if not naming_files:
                print("⚠️ No naming_convention*.json files found")
                # Return default structure instead of empty dict
                return {
                    "message_flows": [],
                    "transformation_requirements": [],
                    "integration_endpoints": [],
                    "database_lookups": [],
                    "business_entities": [],
                    "ace_library_indicators": [],
                    "processing_patterns": [],
                    "technical_specifications": [],
                    "data_enrichment_rules": [],
                    "routing_logic": [],
                    "messageflows": []  # Add our new structure key too
                }
                
            # Step 2: Initialize business requirements structure
            business_requirements = {
                "extraction_timestamp": datetime.now().isoformat(),
                "message_flows": [],
                "transformation_requirements": [],
                "integration_endpoints": [],
                "database_lookups": [],
                "business_entities": [],
                "ace_library_indicators": [],
                "processing_patterns": [],
                "technical_specifications": [],
                "data_enrichment_rules": [],
                "routing_logic": [],
                "messageflows": []
            }
            
            # Step 3: Process each application/flow
            for naming_file in naming_files:
                try:
                    flow_name = naming_file.get('project_naming', {}).get('message_flow_name', '')
                    app_name = naming_file.get('project_naming', {}).get('ace_application_name', '')
                    
                    if not flow_name or not app_name:
                        print(f"⚠️ Missing flow_name or app_name in naming file")
                        continue
                        
                    print(f"\n🔧 Processing Flow: {flow_name}")
                    
                    # Add to old format lists for compatibility
                    business_requirements["message_flows"].append(flow_name)
                    business_requirements["ace_library_indicators"].append(app_name)
                    
                    # Get basic flow metadata
                    input_type = naming_file.get('project_naming', {}).get('input_type', 'MQ')
                    server_name = naming_file.get('project_naming', {}).get('ace_server', 'group-default-server')
                    is_sat_flow = flow_name.endswith("RECSAT") or flow_name.endswith("SNDSAT")
                    
                    # SAFE APPROACH: Use rule-based detection first
                    flow_pattern = {
                        'input_type': input_type,
                        'has_enrichment': False,
                        'has_xsl_transform': True,
                        'xsl_files': [],
                        'has_soap_request': False,
                        'is_synchronous': input_type == "HTTP",
                        'flow_type': "SAT" if is_sat_flow else "RTS" if input_type == "HTTP" else "P2P",
                        'has_event_nodes': True,
                        'has_method_routing': input_type == "HTTP",
                        'routing_methods': []
                    }
                    
                    # Try to enhance with vector DB info if available
                    try:
                        vector_results = safe_vector_query(flow_name)
                        if vector_results and isinstance(vector_results, list):
                            detected_pattern = self.detect_flow_pattern(vector_results)
                            # Merge only if valid
                            if isinstance(detected_pattern, dict):
                                flow_pattern.update(detected_pattern)
                    except Exception as e:
                        print(f"  ⚠️ Vector pattern detection failed: {str(e)}")
                    
                    # Create messageflow requirements entry - FIXED to contain actual values, not placeholders
                    flow_requirements = {
                        "flow_name": flow_name,
                        "app_name": app_name,
                        "server_name": server_name,
                        "input_type": input_type,
                        "is_sat_flow": is_sat_flow,
                        "requires_routing": flow_pattern.get('has_method_routing', False),
                        "routing_methods": flow_pattern.get('routing_methods', []),
                        "flow_type": flow_pattern.get('flow_type'),
                        "has_enrichment": flow_pattern.get('has_enrichment', False),
                        "has_xsl_transform": flow_pattern.get('has_xsl_transform', False),
                        "has_soap_request": flow_pattern.get('has_soap_request', False),
                        "xsl_files": flow_pattern.get('xsl_files', []),
                        "is_synchronous": flow_pattern.get('is_synchronous', False),
                        "has_event_nodes": flow_pattern.get('has_event_nodes', False),
                        "node_connections": [],
                        "technical_description": naming_file.get('business_context', {}).get('description', 
                                                f"Message flow {flow_name} processes data between systems.")
                    }
                    
                    # Add input configuration section
                    flow_requirements["input_configuration"] = {
                        "input_type": input_type,
                        "input_node_name": f"{input_type}Input",
                        "input_details": {
                            "uri_pattern": f"/{flow_name}" if input_type == "HTTP" else "",
                            "http_method": "POST" if input_type == "HTTP" else "",
                            "content_type": "application/xml" if input_type == "HTTP" else "",
                            "queue_name": f"INPUT.{flow_name}.QUEUE" if input_type == "MQ" else "",
                            "file_pattern": f"*.xml" if input_type == "File" else ""
                        }
                    }
                    
                    # Create basic nodes list
                    nodes = []
                    node_connections = []
                    
                    # Entry node based on input type
                    if input_type == "HTTP":
                        nodes.append({
                            "node_id": "FCMComposite_1_1",
                            "node_name": "HTTPInput",
                            "node_type": "HTTPInput",
                            "is_entry_point": True,
                            "node_purpose": "Receives HTTP requests"
                        })
                    elif input_type == "MQ":
                        nodes.append({
                            "node_id": "FCMComposite_1_1",
                            "node_name": "MQInput",
                            "node_type": "MQInput",
                            "is_entry_point": True,
                            "node_purpose": "Receives messages from queue"
                        })
                    elif input_type == "File":
                        nodes.append({
                            "node_id": "FCMComposite_1_1",
                            "node_name": "FileInput",
                            "node_type": "FileInput",
                            "is_entry_point": True,
                            "node_purpose": "Receives files from directory"
                        })
                    
                    # Always add compute node
                    nodes.append({
                        "node_id": "FCMComposite_1_2",
                        "node_name": "ComputeNode",
                        "node_type": "ComputeNode",
                        "is_entry_point": False,
                        "node_purpose": "Main processing node"
                    })
                    
                    # Add routing node if needed
                    if flow_pattern.get('has_method_routing', False):
                        nodes.append({
                            "node_id": "FCMComposite_1_3",
                            "node_name": "RouteNode",
                            "node_type": "RouteNode",
                            "is_entry_point": False,
                            "node_purpose": "Routes based on message content"
                        })
                    
                    # Add output node based on flow type
                    if input_type == "HTTP" and flow_pattern.get('is_synchronous', False):
                        nodes.append({
                            "node_id": "FCMComposite_1_4",
                            "node_name": "HTTPReply",
                            "node_type": "HTTPReply",
                            "is_entry_point": False,
                            "node_purpose": "Sends HTTP reply"
                        })
                    else:
                        nodes.append({
                            "node_id": "FCMComposite_1_4",
                            "node_name": "MQOutput",
                            "node_type": "MQOutput",
                            "is_entry_point": False,
                            "node_purpose": "Sends to output queue"
                        })
                    
                    # Create basic node connections
                    for i in range(len(nodes) - 1):
                        node_connections.append({
                            "source_node_id": nodes[i]["node_id"],
                            "target_node_id": nodes[i + 1]["node_id"],
                            "terminal_type": "out",
                            "filter_condition": None
                        })
                    
                    flow_requirements["nodes"] = nodes
                    flow_requirements["node_connections"] = node_connections
                    
                    business_requirements["messageflows"].append(flow_requirements)
                    
                    # Add queue to endpoints list for old format compatibility
                    if input_type == "MQ":
                        business_requirements["integration_endpoints"].append(f"INPUT.{flow_name}.QUEUE")
                    
                except Exception as e:
                    print(f"⚠️ Error processing {flow_name}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
                        
            # Step 4: Save the enhanced business_requirements.json file
            output_dir = "."  # Use current directory as default
            output_path = Path(output_dir) / "business_requirements.json"
            
            # Ensure directory exists
            try:
                os.makedirs(output_dir, exist_ok=True)
                
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(business_requirements, f, indent=2)
                    
                print(f"✅ Created enhanced business_requirements.json with {len(business_requirements['messageflows'])} flows")
            except Exception as e:
                print(f"⚠️ Error saving JSON file: {e}")
            
            return business_requirements
            
        except Exception as e:
            print(f"❌ Failed to extract business requirements: {e}")
            import traceback
            traceback.print_exc()
            
            # Return the default structure for both formats
            return {
                "message_flows": [],
                "transformation_requirements": [],
                "integration_endpoints": [],
                "database_lookups": [],
                "business_entities": [],
                "ace_library_indicators": [],
                "processing_patterns": [],
                "technical_specifications": [],
                "data_enrichment_rules": [],
                "routing_logic": [],
                "messageflows": [],
                "extraction_timestamp": datetime.now().isoformat()
            }

    def _generate_flow_nodes(self, flow_name, input_type, flow_pattern):
        """Generate node list based on flow characteristics"""
        nodes = []
        
        # Entry node based on input type
        if input_type == "HTTP":
            nodes.append({
                "node_id": "FCMComposite_1_1",
                "node_name": "HTTPInput",
                "node_type": "HTTPInput",
                "is_entry_point": True,
                "node_purpose": "Receives HTTP requests"
            })
        elif input_type == "MQ":
            nodes.append({
                "node_id": "FCMComposite_1_1",
                "node_name": "MQInput",
                "node_type": "MQInput",
                "is_entry_point": True,
                "node_purpose": "Receives messages from queue"
            })
        elif input_type == "File":
            nodes.append({
                "node_id": "FCMComposite_1_1",
                "node_name": "FileInput",
                "node_type": "FileInput",
                "is_entry_point": True,
                "node_purpose": "Receives files from directory"
            })
        
        # Always add start event node
        nodes.append({
            "node_id": "FCMComposite_1_2",
            "node_name": "StartEvent",
            "node_type": "ComputeNode",
            "is_entry_point": False,
            "node_purpose": "Prepares Start Event + Gets Main Identifier"
        })
        
        # If routing is needed, add a router node
        if flow_pattern.get('has_method_routing', False):
            nodes.append({
                "node_id": "FCMComposite_1_3",
                "node_name": "MethodRouter",
                "node_type": "RouteNode",
                "is_entry_point": False,
                "node_purpose": "Routes messages based on method type"
            })
        
        # Add transformation nodes if needed
        if flow_pattern.get('has_xsl_transform', False):
            nodes.append({
                "node_id": "FCMComposite_1_4",
                "node_name": "XSLTransform",
                "node_type": "XSLTransformNode",
                "is_entry_point": False,
                "node_purpose": "Transforms message format"
            })
        
        # Add enrichment nodes if needed
        if flow_pattern.get('has_enrichment', False):
            nodes.append({
                "node_id": "FCMComposite_1_5",
                "node_name": "Enrichment",
                "node_type": "ComputeNode",
                "is_entry_point": False,
                "node_purpose": "Enriches message with additional data"
            })
        
        # Add SOAP request node if needed
        if flow_pattern.get('has_soap_request', False):
            nodes.append({
                "node_id": "FCMComposite_1_6",
                "node_name": "SOAPRequest",
                "node_type": "SOAPRequestNode",
                "is_entry_point": False,
                "node_purpose": "Sends SOAP request to external service"
            })
        
        # Add appropriate output node
        if input_type == "HTTP" and flow_pattern.get('is_synchronous', False):
            nodes.append({
                "node_id": "FCMComposite_1_7",
                "node_name": "HTTPReply",
                "node_type": "HTTPReply",
                "is_entry_point": False,
                "node_purpose": "Returns HTTP response"
            })
        else:
            nodes.append({
                "node_id": "FCMComposite_1_7",
                "node_name": "MQOutput",
                "node_type": "MQOutput",
                "is_entry_point": False,
                "node_purpose": "Sends message to output queue"
            })
        
        # Always add end event node
        nodes.append({
            "node_id": "FCMComposite_1_8",
            "node_name": "EndEvent",
            "node_type": "ComputeNode",
            "is_entry_point": False,
            "node_purpose": "Records end event"
        })
        
        return nodes

    def _generate_node_connections(self, nodes, flow_pattern):
        """Generate connections between nodes"""
        connections = []
        
        # Create a simple linear flow for most nodes
        for i in range(len(nodes) - 1):
            source = nodes[i]
            target = nodes[i + 1]
            
            # Skip connection if source is Router (handled separately)
            if source["node_type"] == "RouteNode":
                continue
                
            connections.append({
                "source_node_id": source["node_id"],
                "target_node_id": target["node_id"],
                "terminal_type": "out",
                "filter_condition": None
            })
        
        # Special handling for routing node if present
        router_node = next((n for n in nodes if n["node_type"] == "RouteNode"), None)
        if router_node:
            # Find target nodes for router
            router_index = next((i for i, n in enumerate(nodes) if n["node_type"] == "RouteNode"), -1)
            if router_index >= 0 and router_index + 1 < len(nodes):
                target = nodes[router_index + 1]
                
                # Add routing methods if available
                routing_methods = flow_pattern.get('routing_methods', [])
                if not routing_methods:
                    routing_methods = ["default"]
                    
                for method in routing_methods:
                    connections.append({
                        "source_node_id": router_node["node_id"],
                        "target_node_id": target["node_id"],
                        "terminal_type": method.lower(),
                        "filter_condition": f"method='{method}'" if method != "default" else None
                    })
        
        return connections

    def _extract_connected_systems(self, pdf_content):
        """Extract connected systems information from PDF content"""
        systems = []
        
        # Try to extract systems from common patterns in PDFs
        system_patterns = [
            r'Connected System\(s\)\s*([A-Za-z0-9_\-\s,]+)',
            r'Source System\s*:\s*([A-Za-z0-9_\-\s]+)',
            r'Target System\s*:\s*([A-Za-z0-9_\-\s]+)'
        ]
        
        extracted_systems = []
        for pattern in system_patterns:
            matches = re.findall(pattern, pdf_content, re.IGNORECASE)
            for match in matches:
                # Split by commas, new lines, or spaces followed by dash
                systems_list = re.split(r'[,\n]|\s+-', match)
                for system in systems_list:
                    system_name = system.strip()
                    if system_name and system_name not in extracted_systems:
                        extracted_systems.append(system_name)
        
        # Try to extract system IDs
        system_id_pattern = r'([A-Z]+-\d+)'
        system_ids = re.findall(system_id_pattern, pdf_content)
        
        # Create system objects
        for i, system_name in enumerate(extracted_systems):
            system_obj = {
                "system_name": system_name,
                "system_id": system_ids[i] if i < len(system_ids) else f"SYS-{i+1:04d}",
                "system_type": self._guess_system_type(system_name),
                "service_endpoints": []
            }
            systems.append(system_obj)
        
        # If no systems extracted, add placeholder
        if not systems:
            systems.append({
                "system_name": "Unknown",
                "system_id": "SYS-0001",
                "system_type": "External",
                "service_endpoints": []
            })
        
        return systems

    def _guess_system_type(self, system_name):
        """Guess system type based on name"""
        system_name = system_name.lower()
        
        if any(db in system_name for db in ['sql', 'oracle', 'db', 'database']):
            return "Database"
        elif any(erp in system_name for erp in ['sap', 'oracle', 'dynamics']):
            return "ERP"
        elif any(mq in system_name for mq in ['mq', 'rabbit', 'kafka']):
            return "Messaging"
        elif any(web in system_name for web in ['http', 'web', 'rest', 'soap']):
            return "WebService"
        elif any(file in system_name for file in ['file', 'ftp', 'sftp']):
            return "File"
        else:
            return "External Service"
    


    def _query_vector_db(self, query: str) -> List[Dict]:
        """Query vector DB for flow-specific information with robust error handling"""
        try:
            # Check if we have a vector store
            if not hasattr(self, 'vector_store') or not self.vector_store:
                try:
                    from vector_knowledge.vector_store import ChromaVectorStore
                    self.vector_store = ChromaVectorStore()
                    print("📋 Created new vector store connection")
                except ImportError:
                    # Try alternative import path
                    try:
                        from vector_knowledge.vector_store import ChromaVectorStore
                        self.vector_store = ChromaVectorStore()
                        print("📋 Created new vector store connection (alternative path)")
                    except Exception as e:
                        print(f"  ⚠️ Could not initialize vector store: {e}")
                        return []
            
            # Query for flow information
            enhanced_query = f"{query} message flow routing structure diagram"
            
            # CRITICAL FIX: Handle vector store results properly
            try:
                # Try the more specific search if available
                if hasattr(self.vector_store, 'search_for_agent_with_diagrams'):
                    try:
                        results = self.vector_store.search_for_agent_with_diagrams(
                            "messageflow_generator", [enhanced_query], top_k=5
                        )
                        
                        # SAFE CHECK: Ensure results are in the expected format
                        if isinstance(results, str):
                            print(f"  ⚠️ Unexpected string result from vector store: {results[:100]}...")
                            return []
                            
                        if results:
                            return results
                    except Exception as e:
                        print(f"  ⚠️ Agent search failed: {e}")
                
                # Fall back to standard search if available
                if hasattr(self.vector_store, 'collection'):
                    results = self.vector_store.collection.query(
                        query_texts=[enhanced_query],
                        n_results=5
                    )
                    
                    # Convert to consistent format
                    formatted_results = []
                    if results and 'documents' in results and results['documents']:
                        for i in range(len(results['documents'][0])):
                            formatted_results.append({
                                'content': results['documents'][0][i],
                                'metadata': results['metadatas'][0][i] if 'metadatas' in results and results['metadatas'][0] else {},
                                'id': results['ids'][0][i] if 'ids' in results and results['ids'][0] else f"result_{i}"
                            })
                        
                    return formatted_results
            except Exception as e:
                print(f"  ⚠️ Vector DB search failed: {e}")
            
            return []
            
        except Exception as e:
            print(f"  ⚠️ Vector DB query failed: {e}")
            return []

    def _extract_node_connections(self, flow_name: str) -> List[Dict]:
        """Extract node connections with robust fallbacks"""
        connections = []
        
        try:
            # Query for technical diagrams and flow architecture
            if hasattr(self, 'vector_store') and self.vector_store:
                query = f"{flow_name} technical diagram node connections architecture"
                results = self._query_vector_db(query)
                
                if results:
                    # Look for content with node connections
                    for result in results:
                        content = result.get('content', '')
                        
                        # Pattern matching for connections in content
                        connection_patterns = [
                            r'(\w+)\s*->\s*(\w+)',  # Node1 -> Node2
                            r'(\w+)\s*connects to\s*(\w+)',  # Node1 connects to Node2
                            r'from\s*(\w+)\s*to\s*(\w+)',  # from Node1 to Node2
                            r'(\w+)\s*node\s*.*?\s*followed by\s*(\w+)'  # Node1 node ... followed by Node2
                        ]
                        
                        import re
                        for pattern in connection_patterns:
                            matches = re.findall(pattern, content, re.IGNORECASE)
                            for match in matches:
                                connections.append({
                                    "from": match[0].strip(),
                                    "to": match[1].strip()
                                })
                    
                    # Use LLM for extraction if available and no connections found
                    if not connections and hasattr(self, 'groq_client') and self.groq_client:
                        most_relevant = results[0].get('content', '')
                        
                        prompt = f"""
                        Extract node connections from this technical diagram description for {flow_name}:
                        
                        {most_relevant[:2000]}
                        
                        Return ONLY a JSON array of connections with 'from' and 'to' fields. 
                        Example:
                        [
                        {{"from": "HTTPInput", "to": "Compute1"}},
                        {{"from": "Compute1", "to": "MQOutput"}}
                        ]
                        """
                        
                        try:
                            response = self.groq_client.chat.completions.create(
                                model=self.groq_model,
                                messages=[{"role": "user", "content": prompt}],
                                temperature=0.1,
                                max_tokens=1000,
                                response_format={"type": "json_object"}
                            )
                            
                            result = response.choices[0].message.content.strip()
                            import json
                            import re
                            
                            # Try to extract JSON array from response
                            json_match = re.search(r'\[\s*\{.*\}\s*\]', result, re.DOTALL)
                            if json_match:
                                connections = json.loads(json_match.group(0))
                            
                        except Exception:
                            pass
        except Exception:
            pass
            
        # If no connections found, generate default linear connections
        if not connections:
            default_nodes = ["Input", "prepareStart", "Compute", "XSLTransform", "Output"]
            for i in range(len(default_nodes) - 1):
                connections.append({
                    "from": default_nodes[i],
                    "to": default_nodes[i+1]
                })
        
        # Remove duplicates
        unique_connections = []
        seen = set()
        for conn in connections:
            key = f"{conn['from']}-{conn['to']}"
            if key not in seen:
                seen.add(key)
                unique_connections.append(conn)
        
        return unique_connections

    def _extract_technical_description(self, flow_name: str) -> str:
        """Extract technical description with robust fallbacks"""
        try:
            if hasattr(self, 'vector_store') and self.vector_store:
                query = f"{flow_name} technical description specification requirements"
                results = self._query_vector_db(query)
                
                if results and results[0].get('content'):
                    most_relevant = results[0].get('content', '')
                    
                    # If LLM available and description is long, summarize
                    if len(most_relevant) > 1000 and hasattr(self, 'groq_client') and self.groq_client:
                        prompt = f"""
                        Summarize this technical description for {flow_name}:
                        
                        {most_relevant[:3000]}
                        
                        Focus on key requirements, integration points, and flow behavior.
                        """
                        
                        try:
                            response = self.groq_client.chat.completions.create(
                                model=self.groq_model,
                                messages=[{"role": "user", "content": prompt}],
                                temperature=0.1,
                                max_tokens=1000
                            )
                            
                            return response.choices[0].message.content.strip()
                            
                        except Exception:
                            # Fall back to original content
                            pass
                    
                    return most_relevant[:1000] + "..." if len(most_relevant) > 1000 else most_relevant
        except Exception:
            pass
        
        # Default description if none found
        return f"Message flow {flow_name} processes data between systems."

    def _discover_naming_conventions(self) -> List[Dict]:
        """Discover naming conventions from files with improved logging"""
        naming_files = []
        
        print("  Looking for naming convention files...")
        
        # Check for single file
        single_file = Path("naming_convention.json")
        if single_file.exists():
            try:
                with open(single_file, 'r') as f:
                    naming_data = json.load(f)
                    naming_files.append(naming_data)
                    flow_name = naming_data.get('project_naming', {}).get('message_flow_name', 'Unknown')
                    print(f"  ✓ Found naming file: {single_file} (flow: {flow_name})")
            except Exception as e:
                print(f"  ⚠️ Error reading {single_file}: {e}")
        
        # Check for numbered files
        idx = 1
        while True:
            numbered_file = Path(f"naming_convention_{idx}.json")
            if numbered_file.exists():
                try:
                    with open(numbered_file, 'r') as f:
                        naming_data = json.load(f)
                        naming_files.append(naming_data)
                        flow_name = naming_data.get('project_naming', {}).get('message_flow_name', 'Unknown')
                        print(f"  ✓ Found naming file: {numbered_file} (flow: {flow_name})")
                except Exception as e:
                    print(f"  ⚠️ Error reading {numbered_file}: {e}")
                idx += 1
            else:
                break
        
        return naming_files

            
    def _extract_technical_diagram(self, flow_name: str) -> Dict:
        """Extract technical diagram information for a flow from vector DB"""
        try:
            # Use the vector DB to find technical diagrams
            query = f"{flow_name} technical diagram architecture"
            results = self.vector_store.search_for_agent_with_diagrams("messageflow_generator", [query], top_k=3)
            
            if not results:
                return None
                
            # Find results with diagrams
            diagram_results = [r for r in results if r.get('metadata', {}).get('has_technical_diagrams')]
            if not diagram_results:
                return None
                
            # Get the best diagram
            best_diagram = diagram_results[0]
            
            # Return diagram information
            return {
                "diagram_content": best_diagram.get('content', ''),
                "ocr_text": best_diagram.get('metadata', {}).get('ocr_text', ''),
                "diagram_score": best_diagram.get('score', 0.0)
            }
        except Exception as e:
            print(f"⚠️ Error extracting technical diagram: {e}")
            return None
        

            
    def extract_node_connections(self, flow_name: str) -> List[Dict]:
        """Extract node connections from technical diagrams or descriptions"""
        try:
            # Query for technical diagrams and flow architecture
            query = f"{flow_name} technical diagram node connections architecture"
            
            # Get results from vector store
            results = self._query_vector_db(query)
            if not results:
                return []
                
            # Look for content with node connections
            connections = []
            
            for result in results:
                content = result.get('content', '')
                
                # Pattern 1: Look for explicit connection descriptions
                connection_patterns = [
                    r'(\w+)\s*->\s*(\w+)',  # Node1 -> Node2
                    r'(\w+)\s*connects to\s*(\w+)',  # Node1 connects to Node2
                    r'from\s*(\w+)\s*to\s*(\w+)',  # from Node1 to Node2
                    r'(\w+)\s*node\s*.*?\s*followed by\s*(\w+)'  # Node1 node ... followed by Node2
                ]
                
                import re
                for pattern in connection_patterns:
                    matches = re.findall(pattern, content, re.IGNORECASE)
                    for match in matches:
                        connections.append({
                            "from": match[0].strip(),
                            "to": match[1].strip()
                        })
            
            # If we couldn't extract connections from patterns, use LLM
            if not connections and self.groq_client:
                # Find the most relevant content
                most_relevant = results[0].get('content', '') if results else ""
                
                prompt = f"""
                Extract node connections from this technical diagram description for {flow_name}:
                
                {most_relevant[:2000]}
                
                Return ONLY a JSON array of connections with 'from' and 'to' fields. 
                Example:
                [
                {{"from": "HTTPInput", "to": "Compute1"}},
                {{"from": "Compute1", "to": "MQOutput"}}
                ]
                
                If you can't determine connections, return an empty array [].
                """
                
                try:
                    response = self.groq_client.chat.completions.create(
                        model=self.groq_model,
                        messages=[{"role": "user", "content": prompt}],
                        temperature=0.1,
                        max_tokens=1000,
                        response_format={"type": "json_object"}
                    )
                    
                    result = response.choices[0].message.content.strip()
                    import json
                    import re
                    
                    # Try to extract JSON array from response
                    json_match = re.search(r'\[\s*\{.*\}\s*\]', result, re.DOTALL)
                    if json_match:
                        connections = json.loads(json_match.group(0))
                    
                except Exception as e:
                    print(f"⚠️ Error extracting connections with LLM: {e}")
            
            # Remove duplicates
            unique_connections = []
            seen = set()
            for conn in connections:
                key = f"{conn['from']}-{conn['to']}"
                if key not in seen:
                    seen.add(key)
                    unique_connections.append(conn)
            
            return unique_connections
                
        except Exception as e:
            print(f"⚠️ Error extracting node connections: {e}")
            return []



    def extract_technical_description(self, flow_name: str) -> str:
        """Extract technical description specific to a flow"""
        try:
            # Query for technical description
            query = f"{flow_name} technical description specification requirements"
            
            # Get results from vector store
            results = self._query_vector_db(query)
            if not results:
                return "No technical description found"
                
            # Find the most relevant description
            most_relevant = results[0].get('content', '') if results else ""
            
            # If description is too long, summarize with LLM
            if len(most_relevant) > 1000 and self.groq_client:
                prompt = f"""
                Summarize this technical description for {flow_name}:
                
                {most_relevant[:3000]}
                
                Focus on key requirements, integration points, and flow behavior.
                Keep it under 500 words.
                """
                
                try:
                    response = self.groq_client.chat.completions.create(
                        model=self.groq_model,
                        messages=[{"role": "user", "content": prompt}],
                        temperature=0.1,
                        max_tokens=1000
                    )
                    
                    return response.choices[0].message.content.strip()
                    
                except Exception as e:
                    print(f"⚠️ Error summarizing technical description: {e}")
                    # Fall back to first 1000 characters
                    return most_relevant[:1000] + "..."
            
            return most_relevant
                
        except Exception as e:
            print(f"⚠️ Error extracting technical description: {e}")
            return "Error extracting technical description"
    



    
    def customize_esql_template(self, template: str, biztalk_components: List[Dict], business_requirements: Dict) -> str:
        """
        LLM: Customize ESQL template using multi-layered approach:
        - Enhanced prompt (Solution 2)
        - Auto-recovery for missing components (Solution 1) 
        - Increased token limit (5000)
        """
        if not self.groq_client:
            raise Exception("LLM client not available - cannot customize ESQL template")
        
        try:
            # Extract business entities and database operations for intelligent customization
            business_entities = business_requirements.get('business_entities', [])
            database_lookups = business_requirements.get('database_lookups', [])
            transformation_rules = business_requirements.get('transformation_requirements', [])
            
            # Create focused business context for customization
            business_context = {
                'entities': business_entities[:5],  # Top 5 entities
                'database_ops': database_lookups[:3],  # Top 3 database operations
                'transformations': transformation_rules[:3]  # Top 3 transformation rules
            }
            
            # SOLUTION 2: Enhanced Prompt with Stronger Instructions
            prompt = f"""CRITICAL: You are customizing an ESQL template. You MUST preserve ALL structural components.

    CURRENT ESQL TEMPLATE:
    {template}

    BUSINESS CONTEXT FOR CUSTOMIZATION:
    {json.dumps(business_context, indent=2)}

    ðŸš¨ ABSOLUTE REQUIREMENTS - THESE MUST BE PRESERVED EXACTLY:
    1. CREATE COMPUTE MODULE line - NEVER change this line
    2. CREATE FUNCTION Main() RETURNS BOOLEAN - NEVER change this line
    3. All DECLARE statements (episInfo, sourceInfo, targetInfo, integrationInfo, dataInfo) - NEVER change
    4. RETURN TRUE; - NEVER change this line
    5. CREATE PROCEDURE CopyMessageHeaders() - MUST be included EXACTLY as provided in input
    6. CREATE PROCEDURE CopyEntireMessage() - MUST be included EXACTLY as provided in input  
    7. END MODULE; - NEVER change this line

    âš¡ CUSTOMIZATION SCOPE (ONLY these specific values can be modified):
    - dataInfo.mainIdentifier XPath expressions based on business entities
    - customReference field VALUES (not the field structure)
    - customProperty field VALUES (not the field structure)
    - Remove unused customReference/customProperty assignments if not needed

      CUSTOMIZATION INSTRUCTIONS:
    - If business entities include "ShipmentInstruction", keep ShipmentId mapping
    - If database lookups mention specific tables, populate relevant customReference fields
    - Update XPath expressions to match business entity structure from requirements
    - Keep ALL sourceInfo and targetInfo field assignments exactly as they are

      CRITICAL OUTPUT REQUIREMENTS:
    - Return the COMPLETE template including ALL procedures
    - Do not truncate, abbreviate, or remove any procedures
    - Every line from the input template must appear in your output
    - Only modify the specific field VALUES mentioned above

    OUTPUT: Return the complete ESQL template with ONLY the specific variable assignments customized based on business requirements."""

            # LLM Call with increased token limit
            response = self.groq_client.chat.completions.create(
                model=os.getenv('GROQ_MODEL', 'llama-3.1-8b-instant'),
                messages=[
                    {"role": "system", "content": "You are an expert IBM ACE ESQL developer. You MUST preserve all structural components and return the COMPLETE template. NEVER truncate or remove procedures."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.05,  # Very low temperature for precise customization
                max_tokens=5000    #   INCREASED TOKEN LIMIT
            )

            # NEW: Add token tracking
            if 'token_tracker' in st.session_state and hasattr(response, 'usage') and response.usage:
                st.session_state.token_tracker.manual_track(
                    agent="biztalk_mapper",
                    operation="esql_customization",
                    model=os.getenv('GROQ_MODEL', 'llama-3.1-8b-instant'),
                    input_tokens=response.usage.prompt_tokens,
                    output_tokens=response.usage.completion_tokens,
                    flow_name=getattr(self, 'current_flow_name', 'biztalk_flow')
                )
            
            # Get the LLM response
            customized_template = response.choices[0].message.content.strip()
            print(f"ðŸ“¥ LLM returned template with {len(customized_template)} characters")
            
            # SOLUTION 1: Auto-Recovery for Missing Components
            missing_procedures = {
                'CREATE PROCEDURE CopyMessageHeaders()': '''CREATE PROCEDURE CopyMessageHeaders() BEGIN
            DECLARE I INTEGER 1;
            DECLARE J INTEGER;
            SET J = CARDINALITY(InputRoot.*[]);
            WHILE I < J DO
                SET OutputRoot.*[I] = InputRoot.*[I];
                SET I = I + 1;
            END WHILE;
        END;''',
                
                'CREATE PROCEDURE CopyEntireMessage()': '''CREATE PROCEDURE CopyEntireMessage() BEGIN
            SET OutputRoot = InputRoot;
        END;'''
            }
            
            # Check for missing procedures and auto-fix
            template_fixed = False
            for proc_name, proc_code in missing_procedures.items():
                if proc_name not in customized_template:
                    print(f"  Auto-fixing missing procedure: {proc_name}")
                    
                    # Smart insertion - place before END MODULE;
                    if 'END MODULE;' in customized_template:
                        # Insert before the final END MODULE;
                        end_module_pos = customized_template.rfind('END MODULE;')
                        before_end = customized_template[:end_module_pos].rstrip()
                        after_end = customized_template[end_module_pos:]
                        customized_template = f"{before_end}\n\n    {proc_code}\n{after_end}"
                    else:
                        # If no END MODULE found, append at the end
                        customized_template += f"\n\n{proc_code}\nEND MODULE;"
                    
                    template_fixed = True
            
            if template_fixed:
                print("  Auto-fixed missing ESQL procedures")
            
            # Enhanced validation with detailed checking
            essential_components = [
                "CREATE COMPUTE MODULE",
                "CREATE FUNCTION Main()",
                "RETURNS BOOLEAN", 
                "DECLARE episInfo",
                "DECLARE sourceInfo",
                "DECLARE targetInfo", 
                "DECLARE integrationInfo",
                "DECLARE dataInfo",
                "RETURN TRUE;",
                "CREATE PROCEDURE CopyMessageHeaders()",
                "CREATE PROCEDURE CopyEntireMessage()",
                "END MODULE;"
            ]
            
            missing_components = []
            for component in essential_components:
                if component not in customized_template:
                    missing_components.append(component)
            
            # Multi-level fallback strategy
            if missing_components:
                print(f"  Still missing components after auto-fix: {missing_components}")
                
                # Try to recover by merging with original template
                if len(missing_components) <= 2:  # Only minor issues
                    print("  Attempting template merge recovery...")
                    
                    # Extract customizations from LLM response
                    customization_patterns = [
                        r'SET dataInfo\.mainIdentifier.*?;',
                        r'SET dataInfo\.customReference.*?;', 
                        r'SET dataInfo\.customProperty.*?;'
                    ]
                    
                    merged_template = template  # Start with original
                    
                    # Apply LLM customizations to original template
                    for pattern in customization_patterns:
                        import re
                        llm_matches = re.findall(pattern, customized_template, re.MULTILINE)
                        for match in llm_matches:
                            # Replace corresponding line in original template
                            field_name = match.split('=')[0].strip()
                            original_pattern = f"{field_name}.*?;"
                            merged_template = re.sub(original_pattern, match, merged_template)
                    
                    customized_template = merged_template
                    print("  Template merge recovery successful")
                    
                else:  # Major issues - use original template
                    print("  Major issues detected - falling back to original template")
                    return template
            
            # Final validation
            final_missing = []
            for component in essential_components:
                if component not in customized_template:
                    final_missing.append(component)
            
            if final_missing:
                print(f"  Final validation failed - missing: {final_missing}")
                print("  Using original template as final fallback")
                return template
            
            # Success metrics
            print("  ESQL template customization successful!")
            print(f"  Original length: {len(template)} characters")
            print(f"  Customized length: {len(customized_template)} characters")
            print(f"  Change ratio: {((len(customized_template) - len(template)) / len(template) * 100):.1f}%")
            
            # Optional: Show what was customized
            if business_entities:
                print(f"  Customized for entities: {business_entities[:3]}")
            if database_lookups:
                print(f"ðŸ—ƒï¸ Applied database lookups: {len(database_lookups)} items")
            
            return customized_template
            
        except Exception as e:
            print(f"  ESQL customization failed: {e}")
            print("  Falling back to original template")
            return template  # Always return something usable



# Main execution for standalone testing
if __name__ == "__main__":
    import argparse
    from pathlib import Path
    
    parser = argparse.ArgumentParser(description='ACE MessageFlow Template Optimizer (Agent 1)')
    parser.add_argument('--pdf-path', required=True, help='Path to PDF file')
    parser.add_argument('--biztalk-path', required=False, help='Path to BizTalk folder (optional)')
    parser.add_argument('--output-dir', default='output', help='Output directory')
    parser.add_argument('--extract-requirements', action='store_true', 
                       help='Extract business requirements for messageflow correction')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("🚀 Agent 1: MessageFlow Template Optimizer")
    print("=" * 60)
    
    # Create mapper instance
    mapper = BizTalkACEMapper()
    
    # Step 1: Query Vector DB for PDF content
    print("\n📝 Step 1: Analyzing PDF content from Vector DB...")
    vector_db_results = []  # Placeholder - replace with actual vector DB query
    
    # Step 2: Detect XSL and Transco presence
    print("\n🔍 Step 2: Detecting XSL and Transco requirements...")
    has_xsl, has_transco = mapper.detect_flow_pattern(vector_db_results)
    
    print(f"   - XSL Transform needed: {'✅ Yes' if has_xsl else '❌ No'}")
    print(f"   - Transco/Enrichment needed: {'✅ Yes' if has_transco else '❌ No'}")
    
    # Step 3: Generate optimized msgflow template
    print("\n⚙️ Step 3: Generating optimized msgflow_template.xml...")
    template_path = mapper.optimize_msgflow_template(has_xsl, has_transco)
    
    # Step 4: (Optional) Process BizTalk components if path provided
    if args.biztalk_path:
        print("\n📦 Step 4: Processing BizTalk components...")
        
        # Get all BizTalk files from directory
        biztalk_files = list(Path(args.biztalk_path).rglob("*.btproj"))
        biztalk_files += list(Path(args.biztalk_path).rglob("*.odx"))
        biztalk_files += list(Path(args.biztalk_path).rglob("*.btm"))
        
        if biztalk_files:
            try:
                result = mapper.process_mapping(biztalk_files, args.pdf_path, args.output_dir)
                print(f"BizTalk processing: {result}")
            except Exception as e:
                print(f"BizTalk processing warning: {e}")
        else:
            print("No BizTalk files found in provided path")
    else:
        print("\nStep 4: Skipped (No BizTalk path provided - PDF-only mode)")
    
    # NEW Step 5: Extract comprehensive business requirements if requested
    if args.extract_requirements:
        print("\n🔍 Step 5: Extracting comprehensive business requirements...")
        requirements_result = mapper.extract_business_requirements(args.output_dir)
        
        if requirements_result['status'] == 'success':
            print(f"✅ Business requirements extracted: {requirements_result['flow_count']} flows")
            print(f"   File: {requirements_result['file_path']}")
        else:
            print(f"❌ Failed to extract business requirements: {requirements_result['message']}")
    
    # Original Step 5: Save metadata for Agent 2 (now Step 6)
    metadata = {
        'has_xsl': has_xsl,
        'has_transco': has_transco,
        'template_path': template_path,
        'pdf_path': args.pdf_path
    }
    
    Path(args.output_dir).mkdir(parents=True, exist_ok=True)
    metadata_path = f"{args.output_dir}/template_metadata.json"
    with open(metadata_path, 'w') as f:
        json.dump(metadata, f, indent=2)
    
    print(f"\nTemplate optimization complete!")
    print(f"   - Template: {template_path}")
    print(f"   - Metadata: {metadata_path}")
    print("\nReady for Agent 2 (fetch_naming.py + pdf_processor.py)")
    print("=" * 60)