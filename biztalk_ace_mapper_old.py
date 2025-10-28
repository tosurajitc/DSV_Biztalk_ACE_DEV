#!/usr/bin/env python3
"""
BizTalk to ACE Specification-Driven Intelligent Mapper
Focus: Business specification-driven mapping with optimized LLM usage
Author: ACE Migration Expert
"""

import os
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any, Union
import PyPDF2
import groq
from dotenv import load_dotenv
from llm_json_parser import LLMJSONParser
import streamlit as st
import lxml.etree as ET

load_dotenv()

class BizTalkACEMapper:
    """Specification-driven BizTalk to ACE intelligent mapper"""
    
    def __init__(self):
        self.groq_client = None
        self.biztalk_components = []
        self.business_requirements = {}
        self.esql_template = ""
        self.mappings = []
        
        # Initialize LLM client
        self._initialize_llm()
    
    def _initialize_llm(self):
        """Initialize GROQ LLM client"""
        api_key = os.getenv('GROQ_API_KEY')
        if api_key:
            try:
                self.groq_client = groq.Groq(api_key=api_key)
                print("âœ… LLM client initialized")
            except Exception as e:
                print(f"âš ï¸ LLM initialization failed: {e}")
                self.groq_client = None
        else:
            print("âš ï¸ GROQ_API_KEY not found in environment")

    # ========================================
    # RULE-BASED FUNCTIONS (100% Rule-Based)
    # ========================================
    
    def parse_biztalk_components(self, biztalk_files: Any) -> List[Dict]:
        """Parse BizTalk component files from folder scan and extract metadata"""
        components = []
        extensions = {'.btproj', '.odx', '.btm', '.xsl', '.xsd', '.cs'}
        
        # Handle both folder path (string) and file list
        if isinstance(biztalk_files, str):
            # Scan folder recursively (3-4 levels deep)
            folder_path = Path(biztalk_files)
            file_paths = []
            
            for root, dirs, files in os.walk(folder_path):
                # Limit to 4 levels deep
                level = len(Path(root).relative_to(folder_path).parts)
                if level > 4:
                    continue
                
                for file in files:
                    if Path(file).suffix.lower() in extensions:
                        full_path = Path(root) / file
                        file_paths.append(full_path)
            
            print(f"ðŸ“ Scanned folder: Found {len(file_paths)} BizTalk files")
        else:
            # Handle list of file paths
            file_paths = [Path(fp) for fp in biztalk_files]
        
        # Parse each file
        for file_path in file_paths:
            try:
                component = self._parse_single_component(file_path)
                if component:
                    components.append(component)
                    
            except Exception as e:
                print(f"âš ï¸ Failed to parse {file_path.name}: {e}")
                continue
        
        print(f"ðŸ” Parsed {len(components)} BizTalk components")
        return components

    def analyze_pdf_content(self, vector_db_results):
        """Detect XSL and Transco mentions from PDF content"""
        has_xsl = False
        has_transco = False
        
        # Search for keywords in vector DB results
        for result in vector_db_results:
            content = result['content'].lower()
            if 'xsl' in content or 'transformation' in content:
                has_xsl = True
            if 'transco' in content or 'enrichment' in content:
                has_transco = True
        
        return has_xsl, has_transco
    


    def optimize_msgflow_template(self, has_xsl, has_transco, template_path="templates/standard_msgflow_template.xml", output_path="msgflow_template.xml"):
        """
        Generate optimized msgflow_template.xml based on PDF content analysis
        
        Args:
            has_xsl (bool): Whether XSL transformation is mentioned in PDF
            has_transco (bool): Whether Transco/Enrichment is mentioned in PDF
            output_path (str): Path to save optimized template
        
        Returns:
            str: Path to generated template
        """
        
        # Ensure template_path is valid - don't overwrite it if it's a parameter
        if not os.path.exists(template_path):
            # Try common fallback locations
            fallback_paths = [
                "templates/standard_msgflow_template.xml",
                "standard_msgflow_template.xml",
                os.path.join(os.path.dirname(__file__), "templates/standard_msgflow_template.xml"),
                os.path.join(os.getcwd(), "templates/standard_msgflow_template.xml")
            ]
            
            for fallback in fallback_paths:
                if os.path.exists(fallback):
                    template_path = fallback
                    print(f"Found template at: {template_path}")
                    break
            else:
                raise FileNotFoundError(f"MessageFlow template not found at {template_path} or any fallback locations")
        
        # Load standard template as text
        with open(template_path, 'r', encoding='utf-8') as f:
            template_content = f.read()
        
        # Track modifications
        modifications = []
        
        # Process BeforeEnrichment and AfterEnrichment nodes
        if not has_transco:
            print("âš ï¸ No Transco mentioned - Removing BeforeEnrichment & AfterEnrichment")
            modifications.append("Removed BeforeEnrichment and AfterEnrichment nodes")
            
            # Use regular expressions to find and remove the BeforeEnrichment node
            import re
            pattern_before = r'<!-- 2\. BEFORE ENRICHMENT SUBFLOW.*?<translation xmi:type="utility:ConstantString" string="BeforeEnrichment"/>\s*</nodes>'
            template_content = re.sub(pattern_before, '', template_content, flags=re.DOTALL)
            
            # Remove the AfterEnrichment node
            pattern_after = r'<!-- 4\. AFTER ENRICHMENT SUBFLOW.*?<translation xmi:type="utility:ConstantString" string="AfterEnrichment"/>\s*</nodes>'
            template_content = re.sub(pattern_after, '', template_content, flags=re.DOTALL)
            
            # Remove connections involving these nodes
            # Connections to BeforeEnrichment (FCMComposite_1_4)
            pattern_conn_before = r'<!-- Connection \d+:.*?FCMComposite_1_4.*?-->\s*<connections.*?/>'
            template_content = re.sub(pattern_conn_before, '', template_content, flags=re.DOTALL)
            
            # Connections to AfterEnrichment (FCMComposite_1_12)
            pattern_conn_after = r'<!-- Connection \d+:.*?FCMComposite_1_12.*?-->\s*<connections.*?/>'
            template_content = re.sub(pattern_conn_after, '', template_content, flags=re.DOTALL)
            
            # Add new connection from MQInput to Compute
            mqinput_to_compute = '''
        <!-- New Connection: MQInput â†’ Compute (After template optimization) -->
        <connections xmi:type="eflow:FCMConnection" xmi:id="FCMConnection_99" 
                    targetNode="FCMComposite_1_1" sourceNode="FCMComposite_1_7" 
                    sourceTerminalName="OutTerminal.Output" targetTerminalName="InTerminal.in"/>
    '''
            # Insert before the closing </composition> tag
            template_content = template_content.replace('</composition>', mqinput_to_compute + '\n    </composition>')
            modifications.append("Added direct connection from MQInput to Compute")
        
        # Process XSLTransform node
        if not has_xsl:
            print("âš ï¸ No XSL mentioned - Removing XSLTransform node")
            modifications.append("Removed XSLTransform node")
            
            # Remove the XSLTransform node
            import re
            pattern_xsl = r'<!-- 3\. XSL TRANSFORM NODE.*?<translation xmi:type="utility:ConstantString" string="XSLTransform"/>\s*</nodes>'
            template_content = re.sub(pattern_xsl, '', template_content, flags=re.DOTALL)
            
            # Remove connections involving the XSLTransform node
            pattern_conn_xsl = r'<!-- Connection \d+:.*?FCMComposite_1_5.*?-->\s*<connections.*?/>'
            template_content = re.sub(pattern_conn_xsl, '', template_content, flags=re.DOTALL)
            
            # Add new direct connection
            if has_transco:
                # Compute to AfterEnrichment
                compute_to_afterenrich = '''
        <!-- New Connection: Compute â†’ AfterEnrichment (After template optimization) -->
        <connections xmi:type="eflow:FCMConnection" xmi:id="FCMConnection_98" 
                    targetNode="FCMComposite_1_12" sourceNode="FCMComposite_1_1" 
                    sourceTerminalName="OutTerminal.out" targetTerminalName="InTerminal.in"/>
    '''
                template_content = template_content.replace('</composition>', compute_to_afterenrich + '\n    </composition>')
                modifications.append("Added direct connection from Compute to AfterEnrichment")
            else:
                # Compute to MQOutput
                compute_to_mqoutput = '''
        <!-- New Connection: Compute â†’ MQOutput (After template optimization) -->
        <connections xmi:type="eflow:FCMConnection" xmi:id="FCMConnection_97" 
                    targetNode="FCMComposite_1_14" sourceNode="FCMComposite_1_1" 
                    sourceTerminalName="OutTerminal.out" targetTerminalName="InTerminal.Input"/>
    '''
                template_content = template_content.replace('</composition>', compute_to_mqoutput + '\n    </composition>')
                modifications.append("Added direct connection from Compute to MQOutput")
        
        # Handle the simplified flow case
        if not has_xsl and not has_transco:
            print("âœ… Simplified flow: MQInput â†’ Compute â†’ MQOutput â†’ SOAPRequest")
        
        # Save the optimized template
        # Create output directory if it doesn't exist
        output_dir = os.path.dirname(output_path)
        if output_dir:  # Only create directories if there's a path specified
            os.makedirs(output_dir, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(template_content)
        
        print(f"âœ… Optimized template saved: {output_path}")
        for mod in modifications:
            print(f"  - {mod}")
        
        return output_path
    


    def _parse_single_component(self, file_path: Path) -> Optional[Dict]:
        """Parse individual BizTalk component file"""
        extension = file_path.suffix.lower()
        
        component = {
            'name': file_path.stem,
            'file_name': file_path.name,
            'file_type': extension,
            'path': str(file_path),
            'size': file_path.stat().st_size if file_path.exists() else 0
        }
        
        # Rule-based component type classification
        if extension == '.btproj':
            component['type'] = 'BizTalk Project'
            component['complexity'] = 'LOW'
        elif extension == '.odx':
            component['type'] = 'Orchestration'
            component['complexity'] = 'HIGH'
        elif extension == '.btm':
            component['type'] = 'BizTalk Map'
            component['complexity'] = 'MEDIUM'
        elif extension == '.xsl':
            component['type'] = 'XSLT Transform'
            component['complexity'] = 'MEDIUM'
        elif extension == '.xsd':
            component['type'] = 'Schema Definition'
            component['complexity'] = 'LOW'
        elif extension == '.cs':
            component['type'] = 'C# Code'
            component['complexity'] = 'HIGH'
        else:
            component['type'] = 'Other File'
            component['complexity'] = 'UNKNOWN'
        
        return component
    
    def extract_pdf_text(self, pdf_file: str) -> str:
        """Extract text content from PDF file"""
        try:
            with open(pdf_file, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text_content = ""
                
                for page in pdf_reader.pages:
                    text_content += page.extract_text() + "\n"
                
                print(f"ðŸ“„ Extracted {len(text_content)} characters from PDF")
                return text_content.strip()
                
        except Exception as e:
            print(f"âŒ PDF extraction failed: {e}")
            return ""
    
    def load_standard_esql_template(self) -> str:
        """Load the standard ESQL template"""
        template = """
CREATE COMPUTE MODULE _SYSTEM___MSG_TYPE___FLOW_PROCESS___SYSTEM2___FLOW_TYPE__InputEventMessage
	CREATE FUNCTION Main() RETURNS BOOLEAN
	BEGIN
		
		DECLARE episInfo 		REFERENCE TO 	Environment.variables.EventData.episInfo;
		DECLARE sourceInfo 		REFERENCE TO 	Environment.variables.EventData.sourceInfo;
		DECLARE targetInfo 		REFERENCE TO 	Environment.variables.EventData.targetInfo;
		DECLARE integrationInfo REFERENCE TO 	Environment.variables.EventData.integrationInfo;
		DECLARE dataInfo 		REFERENCE TO 	Environment.variables.EventData.dataInfo;
		
		SET sourceInfo.srcAppIdentifier 		= InputRoot.XMLNSC.[<].*:Header.*:Source.*:Identifier; 
		SET sourceInfo.srcEnterpriseCode	 	= InputRoot.XMLNSC.[<].*:Header.*:Source.*:EnterpriseCode;
		SET sourceInfo.srcDivision		 		= InputRoot.XMLNSC.[<].*:Header.*:Source.*:Division;
		SET sourceInfo.srcDepartmentCode 		= InputRoot.XMLNSC.[<].*:Header.*:Source.*:DepartmentCode;
		SET sourceInfo.srcBranchCode 			= InputRoot.XMLNSC.[<].*:Header.*:Source.*:BranchCode;
		SET sourceInfo.srcCountryCode 			= InputRoot.XMLNSC.[<].*:Header.*:Source.*:CountryCode;	
		SET sourceInfo.srcCompanyCode 			= InputRoot.XMLNSC.[<].*:Header.*:Source.*:CompanyCode;
		SET sourceInfo.srcApplicationCode 		= InputRoot.XMLNSC.[<].*:Header.*:Source.*:ApplicationCode;
		
		SET targetInfo.tgtAppIdentifier 		= InputRoot.XMLNSC.[<].*:Header.*:Target.*:Identifier; 	
		SET targetInfo.tgtEnterpriseCode 		= InputRoot.XMLNSC.[<].*:Header.*:Target.*:EnterpriseCode; 
		SET targetInfo.tgtDivision 				= InputRoot.XMLNSC.[<].*:Header.*:Target.*:Division; 
		SET targetInfo.tgtDepartmentCode 		= InputRoot.XMLNSC.[<].*:Header.*:Target.*:DepartmentCode; 
		SET targetInfo.tgtBranchCode 			= InputRoot.XMLNSC.[<].*:Header.*:Target.*:branchCode;
		SET targetInfo.tgtCountryCode 			= InputRoot.XMLNSC.[<].*:Header.*:Target.*:CountryCode;  
		SET targetInfo.tgtCompanyCode 			= InputRoot.XMLNSC.[<].*:Header.*:Target.*:CompanyCode; 
		SET targetInfo.tgtApplicationCode 		= InputRoot.XMLNSC.[<].*:Header.*:Target.*:ApplicationCode; 
	
		SET dataInfo.messageType = InputRoot.XMLNSC.[<].*:Header.*:MessageType;		
		SET dataInfo.dataFormat = 'XML';
		SET dataInfo.mainIdentifier = InputRoot.XMLNSC.[<].*:ShipmentInstruction.*:ShipmentDetails.*:ShipmentId;
		SET dataInfo.customReference1		= ''; 						
		SET dataInfo.customReference1Type	= ''; 	
		SET dataInfo.customReference2		= ''; 	
		SET dataInfo.customReference2Type	= ''; 	
		SET dataInfo.customReference3		= ''; 	
		SET dataInfo.customReference3Type	= ''; 
		SET dataInfo.customReference4		= ''; 
		SET dataInfo.customReference4Type	= '';
		SET dataInfo.customProperty1		= '';
		SET dataInfo.customProperty1Type	= '';
		SET dataInfo.customProperty2		= '';
		SET dataInfo.customProperty2Type	= '';
		SET dataInfo.customProperty3		= '';
		SET dataInfo.customProperty3Type	= '';
		SET dataInfo.customProperty4		= '';
		SET dataInfo.customProperty4Type	= '';
		SET dataInfo.batch = false;
		SET OutputRoot=NULL;

	RETURN TRUE;
	END;
END MODULE;
"""
        self.esql_template = template.strip()
        print("ðŸ“‹ Standard ESQL template loaded")
        return self.esql_template
    
    def validate_esql_syntax(self, esql_content: str) -> Dict:
        """Validate ESQL template syntax"""
        validation = {
            'valid': True,
            'errors': [],
            'warnings': []
        }
        
        # Basic syntax validation
        required_elements = [
            'CREATE COMPUTE MODULE',
            'CREATE FUNCTION Main()',
            'RETURNS BOOLEAN',
            'BEGIN',
            'END;',
            'END MODULE;'
        ]
        
        for element in required_elements:
            if element not in esql_content:
                validation['errors'].append(f"Missing required element: {element}")
                validation['valid'] = False
        
        # Check for forbidden patterns
        forbidden_patterns = ['@', '^esql', '```']
        for pattern in forbidden_patterns:
            if pattern in esql_content:
                validation['errors'].append(f"Forbidden pattern found: {pattern}")
                validation['valid'] = False
        
        print(f"ðŸ” ESQL validation: {'âœ… Valid' if validation['valid'] else 'âŒ Invalid'}")
        return validation
    
    # ========================================
    # LLM-BASED FUNCTIONS (100% LLM-Based)
    # ========================================
    
    def extract_business_requirements(self, pdf_content: str) -> Dict:
        """LLM: Extract detailed business requirements in optimized iterations"""
        if not self.groq_client:
            raise Exception("LLM client not available - cannot process business requirements")
        
        if not pdf_content or len(pdf_content.strip()) < 100:
            raise Exception("PDF content is empty or too short - cannot extract business requirements")
        
        try:
            # Step 1: Analyze PDF content size and plan processing strategy
            cleaned_content = pdf_content.replace('\n\n', '\n').replace('\t', ' ').strip()
            total_chars = len(cleaned_content)
            
            # Calculate optimal chunk size (target ~6000 chars per LLM call)
            max_chunk_size = 6000
            num_chunks = max(1, (total_chars + max_chunk_size - 1) // max_chunk_size)
            actual_chunk_size = total_chars // num_chunks
            
            print(f"ðŸ“Š PDF Analysis: {total_chars} chars â†’ {num_chunks} iterations of ~{actual_chunk_size} chars each")
            
            # Step 2: Process content in planned iterations
            all_requirements = {
                "message_flows": [],
                "transformation_requirements": [],
                "integration_endpoints": [],
                "database_lookups": [],
                "business_entities": [],
                "ace_library_indicators": [],
                "processing_patterns": [],
                "technical_specifications": [],
                "data_enrichment_rules": [],
                "routing_logic": []
            }
            
            # Process each chunk with LLM
            for chunk_idx in range(num_chunks):
                start_pos = chunk_idx * actual_chunk_size
                end_pos = min(start_pos + actual_chunk_size + 500, total_chars)  # 500 char overlap
                chunk_content = cleaned_content[start_pos:end_pos]
                
                print(f"ðŸ§  Processing chunk {chunk_idx + 1}/{num_chunks} ({len(chunk_content)} chars)")
                
                prompt = f"""You are an expert IBM ACE architect analyzing business specification chunk {chunk_idx + 1} of {num_chunks}.

BUSINESS SPECIFICATION CHUNK:
{chunk_content}

Extract SPECIFIC technical details from this chunk only. Focus on concrete names, entities, and requirements.

Return JSON with these categories (add items only if found in THIS chunk):
{{
    "message_flows": ["specific flow names and purposes"],
    "transformation_requirements": ["detailed transformation rules with entity names"],
    "integration_endpoints": ["specific queue names, service endpoints, system names"],
    "database_lookups": ["table names, lookup procedures, database operations"],
    "business_entities": ["concrete entity names like ShipmentInstruction, Company, Document"],
    "ace_library_indicators": ["keywords suggesting EPIS_*, CDM_*, routing libraries"],
    "processing_patterns": ["batch, real-time, orchestration, enrichment patterns"],
    "technical_specifications": ["performance numbers, error handling, security"],
    "data_enrichment_rules": ["specific enrichment logic and validation rules"],
    "routing_logic": ["routing conditions, distribution logic, target systems"]
}}

Extract ONLY what is explicitly mentioned in this chunk. Return empty arrays for categories not found."""

                response = self.groq_client.chat.completions.create(
                    model=os.getenv('GROQ_MODEL', 'llama-3.1-8b-instant'),
                    messages=[
                        {"role": "system", "content": "You are an expert IBM ACE architect specializing in technical requirement extraction."},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.1,
                    max_tokens=1500
                )
                if 'token_tracker' in st.session_state and hasattr(response, 'usage') and response.usage:
                    st.session_state.token_tracker.manual_track(
                        agent="biztalk_mapper",
                        operation="business_requirements_extraction",
                        model=os.getenv('GROQ_MODEL', 'llama-3.1-8b-instant'),
                        input_tokens=response.usage.prompt_tokens,
                        output_tokens=response.usage.completion_tokens,
                        flow_name=getattr(self, 'current_flow_name', 'pdf_processing')
                    )

                self.json_parser = LLMJSONParser(debug=True)
                result = self.json_parser.parse_business_requirements(response.choices[0].message.content)
                chunk_requirements = result

                # Merge chunk results into overall requirements
                for category, items in chunk_requirements.items():
                    if category in all_requirements and isinstance(items, list):
                        all_requirements[category].extend(items)
            
            # Remove duplicates and validate final results
            for category in all_requirements:
                all_requirements[category] = list(set(all_requirements[category]))
            
            total_extracted_items = sum(len(v) for v in all_requirements.values())
            
            if total_extracted_items < 5:
                raise Exception(f"Insufficient business requirements extracted - only {total_extracted_items} items found across all chunks")
            
            print(f"âœ… Successfully extracted {total_extracted_items} specific business requirements from {num_chunks} iterations")
            
            # Debug output for verification
            for category, items in all_requirements.items():
                if items:
                    print(f"  ðŸ“‹ {category}: {len(items)} items - {items[:2]}")  # Show first 2 items
            
            return all_requirements
            
        except Exception as e:
            raise Exception(f"Business requirements extraction failed: {e}")

    def generate_excel_output(self, mappings: List[Dict], output_path: str) -> str:
        """Generate comprehensive Excel file with multiple sheets for detailed component mapping"""
        try:
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # SHEET 1: Component Overview (Main Summary)
                overview_data = []
                for mapping in mappings:
                    overview_data.append({
                        'BizTalk Component': mapping.get('biztalk_component', ''),
                        'Component Type': mapping.get('component_type', ''),
                        'Business Purpose': mapping.get('business_purpose', ''),
                        'Primary ACE Artifact': mapping.get('ace_components', {}).get('primary_artifact', {}).get('name', ''),
                        'Primary Artifact Type': mapping.get('ace_components', {}).get('primary_artifact', {}).get('type', ''),
                        'Implementation Priority': mapping.get('implementation_priority', ''),
                        'Confidence': mapping.get('confidence', 0),
                        'Input Queue': mapping.get('ace_components', {}).get('integration_details', {}).get('input_queue', ''),
                        'Output Endpoint': mapping.get('ace_components', {}).get('integration_details', {}).get('output_endpoint', ''),
                        'Reasoning': mapping.get('reasoning', '')
                    })
                
                overview_df = pd.DataFrame(overview_data)
                overview_df.to_excel(writer, sheet_name='Component Overview', index=False)
                
                # SHEET 2: ACE Artifacts Detail
                artifacts_data = []
                for mapping in mappings:
                    biztalk_comp = mapping.get('biztalk_component', '')
                    
                    # Primary artifact
                    primary = mapping.get('ace_components', {}).get('primary_artifact', {})
                    if primary:
                        artifacts_data.append({
                            'BizTalk Component': biztalk_comp,
                            'Artifact Category': 'Primary',
                            'Artifact Type': primary.get('type', ''),
                            'Artifact Name': primary.get('name', ''),
                            'Description/Purpose': primary.get('description', ''),
                            'Implementation Notes': f"Main {primary.get('type', '')} for {biztalk_comp}"
                        })
                    
                    # Supporting artifacts
                    supporting = mapping.get('ace_components', {}).get('supporting_artifacts', [])
                    for artifact in supporting:
                        artifacts_data.append({
                            'BizTalk Component': biztalk_comp,
                            'Artifact Category': 'Supporting',
                            'Artifact Type': artifact.get('type', ''),
                            'Artifact Name': artifact.get('name', ''),
                            'Description/Purpose': artifact.get('purpose', ''),
                            'Implementation Notes': f"Supporting {artifact.get('type', '')} for business logic"
                        })
                
                artifacts_df = pd.DataFrame(artifacts_data)
                artifacts_df.to_excel(writer, sheet_name='ACE Artifacts', index=False)
                
                # SHEET 3: Database Operations
                database_data = []
                for mapping in mappings:
                    biztalk_comp = mapping.get('biztalk_component', '')
                    db_operations = mapping.get('ace_components', {}).get('integration_details', {}).get('database_operations', [])
                    
                    for db_op in db_operations:
                        database_data.append({
                            'BizTalk Component': biztalk_comp,
                            'Database Operation': db_op,
                            'Operation Type': 'Stored Procedure' if db_op.startswith('sp_') or db_op.startswith('proc_') else 'Database Call',
                            'Purpose': f"Data lookup/enrichment for {biztalk_comp}",
                            'Implementation Priority': mapping.get('implementation_priority', 'medium')
                        })
                
                database_df = pd.DataFrame(database_data)
                if not database_df.empty:
                    database_df.to_excel(writer, sheet_name='Database Operations', index=False)
                
                # SHEET 4: Integration Endpoints
                integration_data = []
                for mapping in mappings:
                    biztalk_comp = mapping.get('biztalk_component', '')
                    integration_details = mapping.get('ace_components', {}).get('integration_details', {})
                    
                    input_queue = integration_details.get('input_queue')
                    output_endpoint = integration_details.get('output_endpoint')
                    transformation_logic = integration_details.get('transformation_logic', '')
                    
                    if input_queue or output_endpoint:
                        integration_data.append({
                            'BizTalk Component': biztalk_comp,
                            'Input Queue': input_queue or 'N/A',
                            'Output Endpoint': output_endpoint or 'N/A',
                            'Transformation Logic': transformation_logic,
                            'Integration Pattern': 'Queue-to-Service' if input_queue and output_endpoint else 'Simple Transform',
                            'Configuration Notes': f"Configure MQ settings for {input_queue}" if input_queue else "Internal transformation only"
                        })
                
                integration_df = pd.DataFrame(integration_data)
                if not integration_df.empty:
                    integration_df.to_excel(writer, sheet_name='Integration Points', index=False)
                
                # SHEET 5: Implementation Roadmap
                roadmap_data = []
                priority_order = {'high': 1, 'medium': 2, 'low': 3}
                
                for mapping in mappings:
                    biztalk_comp = mapping.get('biztalk_component', '')
                    priority = mapping.get('implementation_priority', 'medium')
                    primary_artifact = mapping.get('ace_components', {}).get('primary_artifact', {})
                    supporting_artifacts = mapping.get('ace_components', {}).get('supporting_artifacts', [])
                    
                    total_artifacts = 1 + len(supporting_artifacts)  # Primary + supporting
                    estimated_effort = 'High' if total_artifacts > 3 else 'Medium' if total_artifacts > 1 else 'Low'
                    
                    roadmap_data.append({
                        'Priority Rank': priority_order.get(priority, 2),
                        'BizTalk Component': biztalk_comp,
                        'Implementation Priority': priority.title(),
                        'Primary Deliverable': primary_artifact.get('name', ''),
                        'Total Artifacts': total_artifacts,
                        'Estimated Effort': estimated_effort,
                        'Dependencies': ', '.join([art.get('name', '') for art in supporting_artifacts[:3]]),  # First 3 dependencies
                        'Business Impact': mapping.get('business_purpose', '')[:100] + '...' if len(mapping.get('business_purpose', '')) > 100 else mapping.get('business_purpose', '')
                    })
                
                # Sort by priority
                roadmap_df = pd.DataFrame(roadmap_data)
                roadmap_df = roadmap_df.sort_values('Priority Rank').drop('Priority Rank', axis=1)
                roadmap_df.to_excel(writer, sheet_name='Implementation Roadmap', index=False)
                
                # SHEET 6: Legacy Format (for backward compatibility)
                legacy_data = []
                for mapping in mappings:
                    legacy_data.append({
                        'biztalk_component': mapping.get('biztalk_component', ''),
                        'required_ace_library': f"Component-Level: {mapping.get('ace_components', {}).get('primary_artifact', {}).get('type', 'unknown')}"
                    })
                
                legacy_df = pd.DataFrame(legacy_data)
                legacy_df.to_excel(writer, sheet_name='Component Mapping', index=False)
            
            # Format Excel file for better readability
            self._format_excel_file(output_path)
            
            print(f"ðŸ“Š Enhanced Excel file generated: {output_path}")
            print(f"ðŸ“‹ Component mappings: {len(mappings)}")
            print(f"ðŸ“„ Excel sheets: 6 (Overview, Artifacts, Database, Integration, Roadmap, Legacy)")
            
            return output_path
            
        except Exception as e:
            print(f"âŒ Excel generation failed: {e}")
            raise

    def _format_excel_file(self, file_path: str):
        """Apply formatting to Excel file for better readability"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            workbook = load_workbook(file_path)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Format headers (first row)
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(file_path)
            print(f"âœ… Excel formatting applied")
            
        except Exception as e:
            print(f"âš ï¸ Excel formatting failed (but file still usable): {e}")

    def generate_intelligent_mappings(self, biztalk_components: List[Dict], business_requirements: Dict) -> List[Dict]:
        """Generate intelligent BizTalk to ACE component mappings based on business specifications"""
        
        # Validate prerequisites
        if not self.groq_client:
            raise Exception("LLM client required for component-level mapping. Please configure GROQ_API_KEY.")
        
        if not business_requirements:
            raise Exception("Business requirements required for intelligent mapping. Please provide PDF documentation.")
        
        if not biztalk_components:
            raise Exception("No BizTalk components found for mapping.")
        
        try:
            # Prepare component summary
            component_summary = []
            for comp in biztalk_components[:10]:  # Limit to first 10 components
                component_summary.append({
                    'name': comp.get('name', ''),
                    'type': comp.get('type', ''),
                    'complexity': comp.get('complexity', '')
                })
            
            # Create LLM prompt for component-level mapping
            prompt = f"""As an IBM ACE expert, create detailed component mappings for BizTalk to ACE migration.

    BIZTALK COMPONENTS TO MAP:
    {json.dumps(component_summary, indent=2)}

    BUSINESS REQUIREMENTS CONTEXT:
    {json.dumps(business_requirements, indent=2)[:2000]}

    For each BizTalk component, generate specific ACE artifacts with this structure:

    [
        {{
            "biztalk_component": "exact component name",
            "component_type": "BizTalk component type",
            "business_purpose": "what this component does in business terms",
            "ace_components": {{
                "primary_artifact": {{
                    "type": "message_flow|compute_node|xsl_transform|esql_module",
                    "name": "specific_file_name.msgflow|.esql|.xsl",
                    "description": "purpose of this primary artifact"
                }},
                "supporting_artifacts": [
                    {{
                        "type": "esql_module|xsl_transform|subflow|database_lookup",
                        "name": "specific_file_name.ext",
                        "purpose": "what this supporting artifact does"
                    }}
                ],
                "integration_details": {{
                    "input_queue": "specific_queue_name_if_applicable",
                    "output_endpoint": "service_or_queue_name_if_applicable",
                    "database_operations": ["stored_procedure_names"],
                    "transformation_logic": "brief_description_of_business_rules"
                }}
            }},
            "implementation_priority": "high|medium|low",
            "confidence": 0.85,
            "reasoning": "why this component mapping makes business sense"
        }}
    ]

    REQUIREMENTS:
    1. Generate specific, actionable ACE component names
    2. Include database operations from business requirements
    3. Map integration endpoints (queues, services)
    4. Provide transformation logic based on business context
    5. Return valid JSON structure"""

            # Call LLM
            response = self.groq_client.chat.completions.create(
                model=os.getenv('GROQ_MODEL', 'llama-3.1-8b-instant'),
                messages=[
                    {
                        "role": "system", 
                        "content": "You are an expert IBM ACE architect specializing in detailed component-level BizTalk to ACE migrations. Generate precise, actionable component mappings."
                    },
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,  # Low temperature for consistent output
                max_tokens=4000   # Higher limit for detailed component mappings
            )
            if 'token_tracker' in st.session_state and hasattr(response, 'usage') and response.usage:
                st.session_state.token_tracker.manual_track(
                    agent="biztalk_mapper",
                    operation="component_mapping",
                    model=os.getenv('GROQ_MODEL', 'llama-3.1-8b-instant'),
                    input_tokens=response.usage.prompt_tokens,
                    output_tokens=response.usage.completion_tokens,
                    flow_name=getattr(self, 'current_flow_name', 'component_mapping')
                )
            
            # Parse LLM response
            mappings = self.json_parser.parse_component_mappings(response.choices[0].message.content)
            
            # Validate output quality
            if not mappings:
                raise Exception("LLM generated no valid mappings. Check prompt or parsing logic.")
            
            # Quality check - ensure primary artifacts are specified
            valid_mappings = []
            for mapping in mappings:
                if (mapping.get('ace_components') and 
                    mapping.get('ace_components').get('primary_artifact') and
                    mapping.get('ace_components').get('primary_artifact').get('name')):
                    valid_mappings.append(mapping)
                else:
                    print(f"âš ï¸ Skipping incomplete mapping for: {mapping.get('biztalk_component', 'Unknown')}")
            
            if not valid_mappings:
                raise Exception("No valid component mappings generated. All mappings failed quality validation.")
            
            print(f"âœ… Generated {len(valid_mappings)} intelligent component mappings")
            return valid_mappings
            
        except Exception as e:
            raise Exception(f"Component mapping generation failed: {str(e)}")

    def save_mapping_outputs(self, mappings: List[Dict], output_dir: str) -> Dict[str, str]:
        """Save both JSON and Excel outputs for component mappings"""
        
        try:
            os.makedirs(output_dir, exist_ok=True)
            
            # OUTPUT 1: JSON file for successive programs
            json_file_path = os.path.join(output_dir, "biztalk_ace_component_mapping.json")
            
            # Add metadata to JSON output   
            json_output = {
                "metadata": {
                    "generated_at": datetime.now().isoformat(),
                    "mapping_type": "component_level",
                    "total_components": len(mappings),
                    "generator": "BizTalkACEMapper",
                    "version": "2.0"
                },
                "component_mappings": mappings
            }
            
            with open(json_file_path, 'w', encoding='utf-8') as f:
                json.dump(json_output, f, indent=2, ensure_ascii=False)
            
            print(f"ðŸ“„ JSON file saved: {json_file_path}")
            print(f"   â†’ Input for Programs 3, 4, 5...")
            
            # OUTPUT 2: Excel file for user documentation  
            excel_file_path = os.path.join(output_dir, "biztalk_ace_component_mapping.xlsx")
            excel_path = self.generate_excel_output(mappings, excel_file_path)
            
            print(f"ðŸ“Š Excel file saved: {excel_path}")
            print(f"   â†’ Documentation and user review")
            
            return {
                "json_file": json_file_path,
                "excel_file": excel_path
            }
            
        except Exception as e:
            raise Exception(f"Failed to save mapping outputs: {str(e)}")



    def process_mapping(self, biztalk_files: Union[str, List[str], None], pdf_file: str, output_dir: str) -> Dict:
        """Main processing function with Vector DB integration - Supports PDF-only mode"""
        
        try:
            print("ðŸŽ¯ Starting Specification-Driven Component Mapping")
            print("=" * 60)
            
            # Phase 1: BizTalk Component Analysis (Rule-Based) - OPTIONAL
            print("ðŸ” Phase 1: Analyzing BizTalk components...")
            
            if biztalk_files:
                self.biztalk_components = self.parse_biztalk_components(biztalk_files)
                
                if not self.biztalk_components:
                    print("âš ï¸  No BizTalk components found in provided path")
                    self.biztalk_components = []
                else:
                    print(f"âœ… Found {len(self.biztalk_components)} BizTalk components")
            else:
                print("â­ï¸  No BizTalk path provided - PDF-only mode enabled")
                self.biztalk_components = []
            
            # Phase 2: Business Requirements Extraction - VECTOR DB ONLY
            print("ðŸš€ Phase 2: Processing Vector DB focused content...")
            
            if not pdf_file:
                raise Exception("âŒ Vector DB Error: No focused content received from Vector search")
            
            # Use Vector DB content directly
            print(f"ðŸ“Š Vector content received: {len(pdf_file)} characters")
            self.business_requirements = self.extract_business_requirements(pdf_file)
            
            if not self.business_requirements:
                raise Exception("âŒ Vector DB Error: Failed to extract business requirements from focused content")
            
            print(f"âœ… Vector business requirements processed successfully")
            
            # NEW PHASE: Generate MessageFlow Template based on business requirements
            print("âš™ï¸ Phase 2.5: Generating MessageFlow template...")
            
            # Detect XSL and Transco requirements from business requirements
            has_xsl = any("xsl" in str(item).lower() for items in self.business_requirements.values() for item in items)
            has_transco = any("transco" in str(item).lower() or "enrichment" in str(item).lower() 
                            for items in self.business_requirements.values() for item in items)
            
            print(f"   - XSL Transform needed: {'âœ… Yes' if has_xsl else 'âŒ No'}")
            print(f"   - Transco/Enrichment needed: {'âœ… Yes' if has_transco else 'âŒ No'}")
            
            # Generate MessageFlow template
            msgflow_path = None
            try:
                # Look for the standard template in the root folder first
                standard_template_path = "templates/standard_msgflow_template.xml"
                
                if os.path.exists(standard_template_path):
                    # Save the output to the root folder
                    msgflow_path = self.optimize_msgflow_template(
                        has_xsl, 
                        has_transco, 
                        template_path=standard_template_path,
                        output_path="msgflow_template.xml"  # Save in root folder
                    )
                    print(f"âœ… MessageFlow template generated: {msgflow_path}")
                else:
                    print(f"âš ï¸ Standard MessageFlow template not found at {standard_template_path}")
                    print("âš ï¸ Skipping MessageFlow template generation")
            except Exception as e:
                print(f"âš ï¸ MessageFlow template generation failed: {str(e)}")
            
            # Phase 3: Component-Level Mapping (LLM) - CONDITIONAL
            print("ðŸ§  Phase 3: Generating component mappings...")
            
            # âœ… FIX: Only generate mappings if BizTalk components exist
            if self.biztalk_components:
                self.mappings = self.generate_intelligent_mappings(
                    self.biztalk_components,
                    self.business_requirements  
                )
                
                if not self.mappings:
                    print("âš ï¸  No valid component mappings generated")
                    self.mappings = []
                else:
                    print(f"âœ… Generated {len(self.mappings)} component mappings")
            else:
                print("â­ï¸  Skipping component mappings (No BizTalk components to map)")
                self.mappings = []
            
            # Phase 4: ESQL Template Customization (Optional Enhancement)
            print("ðŸ“ Phase 4: Customizing ESQL template...")
            
            customized_path = None
            # âœ… FIX: Check if template_path exists and is valid
            template_path = "ESQL_Template_Updated.esql"
            
            if template_path and os.path.exists(template_path):
                try:
                    with open(template_path, 'r', encoding='utf-8') as f:
                        self.esql_template = f.read()
                    
                    # Customize template using mappings (works even without BizTalk components)
                    customized_template = self.customize_esql_template(
                        self.esql_template, 
                        self.biztalk_components, 
                        self.business_requirements
                    )
                    
                    # Save customized template
                    customized_path = os.path.join(output_dir, "ESQL_Template_Updated.esql")
                    os.makedirs(output_dir, exist_ok=True)
                    with open(customized_path, 'w', encoding='utf-8') as f:
                        f.write(customized_template)
                    print(f"ðŸ“ Customized ESQL template saved: {customized_path}")
                except Exception as e:
                    print(f"âš ï¸ ESQL template customization failed: {e}")
                    customized_path = None
            else:
                print("âš ï¸ Base ESQL template not found, skipping customization")
            
            # Phase 5: Generate Output Files - CONDITIONAL
            print("ðŸ’¾ Phase 5: Generating output files...")
            
            # Initialize output_files dictionary
            output_files = {"json_file": None, "excel_file": None}
            
            # Only generate mapping outputs if we have mappings
            if self.mappings:
                output_files = self.save_mapping_outputs(self.mappings, output_dir)
                print(f"âœ… Mapping outputs saved")
            else:
                print("â­ï¸  No component mappings to save (PDF-only mode)")
                # Still save business requirements for downstream use
                requirements_path = os.path.join(output_dir, "business_requirements.json")
                os.makedirs(output_dir, exist_ok=True)
                with open(requirements_path, 'w', encoding='utf-8') as f:
                    json.dump(self.business_requirements, f, indent=2, ensure_ascii=False)
                print(f"ðŸ“„ Business requirements saved: {requirements_path}")
                output_files["json_file"] = requirements_path
            
            # Return comprehensive results
            result = {
                "components_processed": len(self.biztalk_components),
                "mappings_generated": len(self.mappings),
                "business_requirements_found": True,
                "json_file": output_files.get("json_file"),
                "excel_file": output_files.get("excel_file"),
                "template_file": customized_path,
                "vector_processing": True,
                "vector_content_length": len(pdf_file),
                "mode": "pdf_only" if not self.biztalk_components else "full_mapping"
            }
            
            # Add MessageFlow template path to results if it was created
            if msgflow_path:
                result["msgflow_template"] = msgflow_path
            
            print("ðŸŽ¯ Specification-driven mapping completed successfully!")
            if not self.biztalk_components:
                print("ðŸ“‹ PDF-only mode: Business requirements extracted, ready for Agent 2")
            else:
                print(f"ðŸ“Š Full mapping mode: {len(self.mappings)} components mapped")
            
            return result
            
        except Exception as e:
            # Let Vector DB errors propagate to UI
            error_message = str(e)
            if "Vector DB Error" in error_message:
                print(f"âŒ {error_message}")
                raise Exception(f"Vector DB Processing Failed: {error_message}")
            else:
                print(f"ðŸ’¥ Error in process_mapping: {error_message}")
                raise Exception(f"Component Mapping Failed: {error_message}")
        


    
    def customize_esql_template(self, template: str, biztalk_components: List[Dict], business_requirements: Dict) -> str:
        """
        LLM: Customize ESQL template using multi-layered approach:
        - Enhanced prompt (Solution 2)
        - Auto-recovery for missing components (Solution 1) 
        - Increased token limit (5000)
        """
        if not self.groq_client:
            raise Exception("LLM client not available - cannot customize ESQL template")
        
        try:
            # Extract business entities and database operations for intelligent customization
            business_entities = business_requirements.get('business_entities', [])
            database_lookups = business_requirements.get('database_lookups', [])
            transformation_rules = business_requirements.get('transformation_requirements', [])
            
            # Create focused business context for customization
            business_context = {
                'entities': business_entities[:5],  # Top 5 entities
                'database_ops': database_lookups[:3],  # Top 3 database operations
                'transformations': transformation_rules[:3]  # Top 3 transformation rules
            }
            
            # SOLUTION 2: Enhanced Prompt with Stronger Instructions
            prompt = f"""CRITICAL: You are customizing an ESQL template. You MUST preserve ALL structural components.

    CURRENT ESQL TEMPLATE:
    {template}

    BUSINESS CONTEXT FOR CUSTOMIZATION:
    {json.dumps(business_context, indent=2)}

    ðŸš¨ ABSOLUTE REQUIREMENTS - THESE MUST BE PRESERVED EXACTLY:
    1. CREATE COMPUTE MODULE line - NEVER change this line
    2. CREATE FUNCTION Main() RETURNS BOOLEAN - NEVER change this line
    3. All DECLARE statements (episInfo, sourceInfo, targetInfo, integrationInfo, dataInfo) - NEVER change
    4. RETURN TRUE; - NEVER change this line
    5. CREATE PROCEDURE CopyMessageHeaders() - MUST be included EXACTLY as provided in input
    6. CREATE PROCEDURE CopyEntireMessage() - MUST be included EXACTLY as provided in input  
    7. END MODULE; - NEVER change this line

    âš¡ CUSTOMIZATION SCOPE (ONLY these specific values can be modified):
    - dataInfo.mainIdentifier XPath expressions based on business entities
    - customReference field VALUES (not the field structure)
    - customProperty field VALUES (not the field structure)
    - Remove unused customReference/customProperty assignments if not needed

    ðŸŽ¯ CUSTOMIZATION INSTRUCTIONS:
    - If business entities include "ShipmentInstruction", keep ShipmentId mapping
    - If database lookups mention specific tables, populate relevant customReference fields
    - Update XPath expressions to match business entity structure from requirements
    - Keep ALL sourceInfo and targetInfo field assignments exactly as they are

    âš ï¸ CRITICAL OUTPUT REQUIREMENTS:
    - Return the COMPLETE template including ALL procedures
    - Do not truncate, abbreviate, or remove any procedures
    - Every line from the input template must appear in your output
    - Only modify the specific field VALUES mentioned above

    OUTPUT: Return the complete ESQL template with ONLY the specific variable assignments customized based on business requirements."""

            # LLM Call with increased token limit
            response = self.groq_client.chat.completions.create(
                model=os.getenv('GROQ_MODEL', 'llama-3.1-8b-instant'),
                messages=[
                    {"role": "system", "content": "You are an expert IBM ACE ESQL developer. You MUST preserve all structural components and return the COMPLETE template. NEVER truncate or remove procedures."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.05,  # Very low temperature for precise customization
                max_tokens=5000    # âœ… INCREASED TOKEN LIMIT
            )

            # NEW: Add token tracking
            if 'token_tracker' in st.session_state and hasattr(response, 'usage') and response.usage:
                st.session_state.token_tracker.manual_track(
                    agent="biztalk_mapper",
                    operation="esql_customization",
                    model=os.getenv('GROQ_MODEL', 'llama-3.1-8b-instant'),
                    input_tokens=response.usage.prompt_tokens,
                    output_tokens=response.usage.completion_tokens,
                    flow_name=getattr(self, 'current_flow_name', 'biztalk_flow')
                )
            
            # Get the LLM response
            customized_template = response.choices[0].message.content.strip()
            print(f"ðŸ“¥ LLM returned template with {len(customized_template)} characters")
            
            # SOLUTION 1: Auto-Recovery for Missing Components
            missing_procedures = {
                'CREATE PROCEDURE CopyMessageHeaders()': '''CREATE PROCEDURE CopyMessageHeaders() BEGIN
            DECLARE I INTEGER 1;
            DECLARE J INTEGER;
            SET J = CARDINALITY(InputRoot.*[]);
            WHILE I < J DO
                SET OutputRoot.*[I] = InputRoot.*[I];
                SET I = I + 1;
            END WHILE;
        END;''',
                
                'CREATE PROCEDURE CopyEntireMessage()': '''CREATE PROCEDURE CopyEntireMessage() BEGIN
            SET OutputRoot = InputRoot;
        END;'''
            }
            
            # Check for missing procedures and auto-fix
            template_fixed = False
            for proc_name, proc_code in missing_procedures.items():
                if proc_name not in customized_template:
                    print(f"âš ï¸ Auto-fixing missing procedure: {proc_name}")
                    
                    # Smart insertion - place before END MODULE;
                    if 'END MODULE;' in customized_template:
                        # Insert before the final END MODULE;
                        end_module_pos = customized_template.rfind('END MODULE;')
                        before_end = customized_template[:end_module_pos].rstrip()
                        after_end = customized_template[end_module_pos:]
                        customized_template = f"{before_end}\n\n    {proc_code}\n{after_end}"
                    else:
                        # If no END MODULE found, append at the end
                        customized_template += f"\n\n{proc_code}\nEND MODULE;"
                    
                    template_fixed = True
            
            if template_fixed:
                print("ðŸ”§ Auto-fixed missing ESQL procedures")
            
            # Enhanced validation with detailed checking
            essential_components = [
                "CREATE COMPUTE MODULE",
                "CREATE FUNCTION Main()",
                "RETURNS BOOLEAN", 
                "DECLARE episInfo",
                "DECLARE sourceInfo",
                "DECLARE targetInfo", 
                "DECLARE integrationInfo",
                "DECLARE dataInfo",
                "RETURN TRUE;",
                "CREATE PROCEDURE CopyMessageHeaders()",
                "CREATE PROCEDURE CopyEntireMessage()",
                "END MODULE;"
            ]
            
            missing_components = []
            for component in essential_components:
                if component not in customized_template:
                    missing_components.append(component)
            
            # Multi-level fallback strategy
            if missing_components:
                print(f"âš ï¸ Still missing components after auto-fix: {missing_components}")
                
                # Try to recover by merging with original template
                if len(missing_components) <= 2:  # Only minor issues
                    print("ðŸ”„ Attempting template merge recovery...")
                    
                    # Extract customizations from LLM response
                    customization_patterns = [
                        r'SET dataInfo\.mainIdentifier.*?;',
                        r'SET dataInfo\.customReference.*?;', 
                        r'SET dataInfo\.customProperty.*?;'
                    ]
                    
                    merged_template = template  # Start with original
                    
                    # Apply LLM customizations to original template
                    for pattern in customization_patterns:
                        import re
                        llm_matches = re.findall(pattern, customized_template, re.MULTILINE)
                        for match in llm_matches:
                            # Replace corresponding line in original template
                            field_name = match.split('=')[0].strip()
                            original_pattern = f"{field_name}.*?;"
                            merged_template = re.sub(original_pattern, match, merged_template)
                    
                    customized_template = merged_template
                    print("âœ… Template merge recovery successful")
                    
                else:  # Major issues - use original template
                    print("ðŸ”„ Major issues detected - falling back to original template")
                    return template
            
            # Final validation
            final_missing = []
            for component in essential_components:
                if component not in customized_template:
                    final_missing.append(component)
            
            if final_missing:
                print(f"âŒ Final validation failed - missing: {final_missing}")
                print("ðŸ”„ Using original template as final fallback")
                return template
            
            # Success metrics
            print("âœ… ESQL template customization successful!")
            print(f"ðŸ“Š Original length: {len(template)} characters")
            print(f"ðŸ“Š Customized length: {len(customized_template)} characters")
            print(f"ðŸ“Š Change ratio: {((len(customized_template) - len(template)) / len(template) * 100):.1f}%")
            
            # Optional: Show what was customized
            if business_entities:
                print(f"ðŸŽ¯ Customized for entities: {business_entities[:3]}")
            if database_lookups:
                print(f"ðŸ—ƒï¸ Applied database lookups: {len(database_lookups)} items")
            
            return customized_template
            
        except Exception as e:
            print(f"âŒ ESQL customization failed: {e}")
            print("ðŸ”„ Falling back to original template")
            return template  # Always return something usable


# Main execution for standalone testing
if __name__ == "__main__":
    import argparse
    from pathlib import Path
    
    parser = argparse.ArgumentParser(description='ACE MessageFlow Template Optimizer (Agent 1)')
    parser.add_argument('--pdf-path', required=True, help='Path to PDF file')
    parser.add_argument('--biztalk-path', required=False, help='Path to BizTalk folder (optional)')
    parser.add_argument('--output-dir', default='output', help='Output directory')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("ðŸš€ Agent 1: MessageFlow Template Optimizer")
    print("=" * 60)
    
    # Create mapper instance
    mapper = BizTalkACEMapper()
    
    # Step 1: Query Vector DB for PDF content
    print("\nðŸ“Š Step 1: Analyzing PDF content from Vector DB...")
    # TODO: Implement query_vector_db() or integrate with existing PDF processor
    vector_db_results = []  # Placeholder - replace with actual vector DB query
    
    # Step 2: Detect XSL and Transco presence
    print("\nðŸ” Step 2: Detecting XSL and Transco requirements...")
    has_xsl, has_transco = mapper.analyze_pdf_content(vector_db_results)
    
    print(f"   - XSL Transform needed: {'âœ… Yes' if has_xsl else 'âŒ No'}")
    print(f"   - Transco/Enrichment needed: {'âœ… Yes' if has_transco else 'âŒ No'}")
    
    # Step 3: Generate optimized msgflow template
    print("\nâš™ï¸  Step 3: Generating optimized msgflow_template.xml...")
    template_path = mapper.optimize_msgflow_template(has_xsl, has_transco)
    
    # Step 4: (Optional) Process BizTalk components if path provided
    if args.biztalk_path:
        print("\nðŸ“¦ Step 4: Processing BizTalk components...")
        
        # Get all BizTalk files from directory
        biztalk_files = list(Path(args.biztalk_path).rglob("*.btproj"))
        biztalk_files += list(Path(args.biztalk_path).rglob("*.odx"))
        biztalk_files += list(Path(args.biztalk_path).rglob("*.btm"))
        
        if biztalk_files:
            try:
                result = mapper.process_mapping(biztalk_files, args.pdf_path, args.output_dir)
                print(f"âœ… BizTalk processing: {result}")
            except Exception as e:
                print(f"âš ï¸  BizTalk processing warning: {e}")
        else:
            print("âš ï¸  No BizTalk files found in provided path")
    else:
        print("\nâ­ï¸  Step 4: Skipped (No BizTalk path provided - PDF-only mode)")
    
    # Step 5: Save metadata for Agent 2
    metadata = {
        'has_xsl': has_xsl,
        'has_transco': has_transco,
        'template_path': template_path,
        'pdf_path': args.pdf_path
    }
    
    Path(args.output_dir).mkdir(parents=True, exist_ok=True)
    metadata_path = f"{args.output_dir}/template_metadata.json"
    with open(metadata_path, 'w') as f:
        json.dump(metadata, f, indent=2)
    
    print(f"\nâœ… Template optimization complete!")
    print(f"   - Template: {template_path}")
    print(f"   - Metadata: {metadata_path}")
    print("\nðŸŽ¯ Ready for Agent 2 (fetch_naming.py + pdf_processor.py)")
    print("=" * 60)